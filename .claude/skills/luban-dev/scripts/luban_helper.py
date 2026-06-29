#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
luban_skill - Luban 配置编辑器辅助脚本
用于 AI 操作 Luban 配置表、枚举、Bean 等
"""

import os
import sys

# Windows 编码修复：必须在最开始执行
if sys.platform == 'win32':
    os.environ['PYTHONIOENCODING'] = 'utf-8'
    import io
    # 强制重新打开 stdout/stderr 使用 UTF-8
    if sys.stdout.encoding != 'utf-8':
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', line_buffering=True)
    if sys.stderr.encoding != 'utf-8':
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', line_buffering=True)

import argparse
import json
import hashlib
from pathlib import Path
from typing import Optional, List, Dict, Any, Tuple
from dataclasses import dataclass, field

try:
    import openpyxl
    from openpyxl import Workbook
except ImportError:
    print("错误: 请先安装 openpyxl: pip install openpyxl")
    sys.exit(1)


@dataclass
class LubanSheetStructure:
    """Luban xlsx 工作表结构解析结果（统一解析器输出）"""
    var_row_num: int = 0                # ##var 行号（1-indexed）
    type_row_num: Optional[int] = None  # ##type 行号（None=缺失）
    group_row_num: Optional[int] = None # ##group 行号（None=缺失）
    comment_row_nums: List[int] = field(default_factory=list)  # ## 注释行号列表
    data_start_row: int = 0             # 数据起始行号（1-indexed）
    fields: List[Dict[str, Any]] = field(default_factory=list) # 标准化字段列表
    var_row: Optional[List] = None      # 原始 ##var 行数据
    type_row: Optional[List] = None     # 原始 ##type 行数据
    group_row: Optional[List] = None    # 原始 ##group 行数据
    format_errors: List[str] = field(default_factory=list)   # 结构错误（阻止 Luban 生成）
    format_warnings: List[str] = field(default_factory=list) # 结构警告（不影响生成但应修复）

    @property
    def is_valid(self) -> bool:
        """结构是否有效（无 format_errors）"""
        return len(self.format_errors) == 0


class LubanConfigHelper:
    """Luban 配置辅助类"""

    @staticmethod
    def _cell(row: tuple, index: int, default: Any = "") -> Any:
        """从 xlsx 行元组安全取值，空单元格(None)规范化为 default。

        openpyxl 对空单元格返回 None，此方法统一处理：
        - 索引越界 → default
        - 值为 None → default
        - 其他值 → 原值
        """
        if index >= len(row):
            return default
        val = row[index]
        return default if val is None else val

    def __init__(self, data_dir: str, cache_dir: str = ".luban_cache"):
        self.data_dir = Path(data_dir)
        self.cache_dir = Path(cache_dir)
        self.cache_dir.mkdir(exist_ok=True)
        
        # 定义文件路径
        self.enums_file = self.data_dir / "__enums__.xlsx"
        self.beans_file = self.data_dir / "__beans__.xlsx"
        self.tables_file = self.data_dir / "__tables__.xlsx"
    
    # ==================== 枚举操作 ====================
    
    def list_enums(self) -> List[Dict[str, Any]]:
        """列出所有枚举"""
        if not self.enums_file.exists():
            return []
        
        wb = openpyxl.load_workbook(self.enums_file)
        sheet = wb.active
        
        enums = []
        current_enum = None
        
        for row in sheet.iter_rows(min_row=4, values_only=True):
            full_name = row[1]  # B列
            
            if full_name:  # 新枚举定义
                if current_enum:
                    enums.append(current_enum)
                current_enum = {
                    "full_name": full_name,
                    "flags": row[2],
                    "unique": row[3],
                    "comment": self._cell(row, 6),
                    "items": []
                }
                # 检查同一行是否有第一个枚举项
                first_item_name = self._cell(row, 7, None)
                if first_item_name:
                    current_enum["items"].append({
                        "name": first_item_name,
                        "alias": self._cell(row, 8),
                        "value": self._cell(row, 9, None),
                        "comment": self._cell(row, 10)
                    })
            elif current_enum:  # 枚举项
                # *items 列开始 (H列开始，索引7)
                item_name = self._cell(row, 7, None)
                if item_name:
                    current_enum["items"].append({
                        "name": item_name,
                        "alias": self._cell(row, 8),
                        "value": self._cell(row, 9, None),
                        "comment": self._cell(row, 10)
                    })
        
        if current_enum:
            enums.append(current_enum)
        
        wb.close()
        return enums
    
    def get_enum(self, enum_name: str) -> Optional[Dict[str, Any]]:
        """获取指定枚举"""
        enums = self.list_enums()
        for enum in enums:
            if enum["full_name"] == enum_name or enum["full_name"].endswith("." + enum_name):
                return enum
        return None
    
    def add_enum(self, full_name: str, items: List[Dict], flags: bool = False, 
                 unique: bool = True, comment: str = "") -> bool:
        """新增枚举"""
        if not self.enums_file.exists():
            print(f"错误: 文件不存在 {self.enums_file}")
            return False
        
        # 检查是否已存在
        existing = self.get_enum(full_name)
        if existing:
            print(f"错误: 枚举 {full_name} 已存在")
            return False
        
        wb = openpyxl.load_workbook(self.enums_file)
        sheet = wb.active
        
        # 找到最后一行
        last_row = sheet.max_row
        
        # 添加枚举定义行
        row_num = last_row + 1
        sheet.cell(row=row_num, column=2, value=full_name)  # full_name
        sheet.cell(row=row_num, column=3, value=flags)       # flags
        sheet.cell(row=row_num, column=4, value=unique)      # unique
        sheet.cell(row=row_num, column=7, value=comment)     # comment
        
        # 添加第一个枚举项
        if items:
            item = items[0]
            sheet.cell(row=row_num, column=8, value=item.get("name"))
            sheet.cell(row=row_num, column=9, value=item.get("alias", ""))
            sheet.cell(row=row_num, column=10, value=item.get("value"))
            sheet.cell(row=row_num, column=11, value=item.get("comment", ""))
        
        # 添加剩余枚举项
        for i, item in enumerate(items[1:], start=1):
            row_num = last_row + 1 + i
            sheet.cell(row=row_num, column=8, value=item.get("name"))
            sheet.cell(row=row_num, column=9, value=item.get("alias", ""))
            sheet.cell(row=row_num, column=10, value=item.get("value"))
            sheet.cell(row=row_num, column=11, value=item.get("comment", ""))
        
        wb.save(self.enums_file)
        wb.close()
        
        print(f"✓ 已添加枚举: {full_name}")
        return True
    
    def delete_enum(self, enum_name: str) -> bool:
        """删除枚举"""
        if not self.enums_file.exists():
            print(f"错误: 文件不存在 {self.enums_file}")
            return False
        
        wb = openpyxl.load_workbook(self.enums_file)
        sheet = wb.active
        
        # 找到枚举的起始和结束行
        start_row = None
        end_row = None
        
        for i, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
            full_name = row[1]
            
            if full_name == enum_name or (full_name and full_name.endswith("." + enum_name)):
                start_row = i
            elif start_row and full_name:  # 遇到下一个枚举
                end_row = i - 1
                break
        
        if not start_row:
            print(f"错误: 未找到枚举 {enum_name}")
            wb.close()
            return False
        
        if not end_row:
            end_row = sheet.max_row
        
        # 从后往前删除行
        for row_num in range(end_row, start_row - 1, -1):
            sheet.delete_rows(row_num)
        
        wb.save(self.enums_file)
        wb.close()
        
        print(f"✓ 已删除枚举: {enum_name}")
        return True
    
    def update_enum(self, enum_name: str, comment: str = None, flags: bool = None) -> bool:
        """更新枚举属性
        
        Args:
            enum_name: 枚举名称
            comment: 注释
            flags: 是否为标志枚举
        """
        if not self.enums_file.exists():
            print(f"错误: 文件不存在 {self.enums_file}")
            return False
        
        wb = openpyxl.load_workbook(self.enums_file)
        sheet = wb.active
        
        # 找到枚举定义行
        target_row = None
        for i, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
            full_name = row[1]  # B列
            if full_name == enum_name or (full_name and full_name.endswith("." + enum_name)):
                target_row = i
                break
        
        if not target_row:
            print(f"错误: 未找到枚举 {enum_name}")
            wb.close()
            return False
        
        # 更新属性
        updated = []
        if comment is not None:
            sheet.cell(row=target_row, column=6, value=comment)  # F列 = comment
            updated.append(f"comment={comment}")
        if flags is not None:
            sheet.cell(row=target_row, column=5, value=flags)  # E列 = flags
            updated.append(f"flags={flags}")
        
        if not updated:
            print("警告: 没有指定要更新的属性")
            wb.close()
            return False
        
        wb.save(self.enums_file)
        wb.close()
        
        print(f"✓ 已更新枚举: {enum_name}")
        print(f"  更新属性: {', '.join(updated)}")
        return True
    
    # ==================== Bean 操作 ====================
    
    def list_beans(self) -> List[Dict[str, Any]]:
        """列出所有 Bean"""
        if not self.beans_file.exists():
            return []
        
        wb = openpyxl.load_workbook(self.beans_file)
        sheet = wb.active
        
        beans = []
        current_bean = None
        
        for row in sheet.iter_rows(min_row=4, values_only=True):
            full_name = row[1]  # B列
            
            if full_name:  # 新 Bean 定义
                if current_bean:
                    beans.append(current_bean)
                current_bean = {
                    "full_name": full_name,
                    "parent": self._cell(row, 2),
                    "value_type": self._cell(row, 3),
                    "sep": self._cell(row, 4),
                    "alias": self._cell(row, 5),
                    "comment": self._cell(row, 6),
                    "group": self._cell(row, 8),
                    "fields": []
                }
                # 检查同一行是否有第一个字段
                first_field_name = self._cell(row, 9, None)
                if first_field_name:
                    current_bean["fields"].append({
                        "name": first_field_name,
                        "alias": self._cell(row, 10),
                        "type": self._cell(row, 11),
                        "group": self._cell(row, 12),
                        "comment": self._cell(row, 13)
                    })
            elif current_bean:  # 字段行
                # *fields 列开始 (I列开始，索引9)
                field_name = self._cell(row, 9, None)
                if field_name:
                    current_bean["fields"].append({
                        "name": field_name,
                        "alias": self._cell(row, 10),
                        "type": self._cell(row, 11),
                        "group": self._cell(row, 12),
                        "comment": self._cell(row, 13)
                    })
        
        if current_bean:
            beans.append(current_bean)
        
        wb.close()
        return beans
    
    def get_bean(self, bean_name: str) -> Optional[Dict[str, Any]]:
        """获取指定 Bean"""
        beans = self.list_beans()
        for bean in beans:
            if bean["full_name"] == bean_name or bean["full_name"].endswith("." + bean_name):
                return bean
        return None
    
    def add_bean(self, full_name: str, fields: List[Dict], parent: str = "",
                 comment: str = "", alias: str = "", value_type: int = 0,
                 sep: str = "") -> bool:
        """新增 Bean

        Args:
            full_name: Bean 全名
            fields: 字段列表
            parent: 父类名称
            comment: 注释
            alias: 别名
            value_type: 是否为值类型（0=普通类，1=值类型/struct），用于 list<Bean> 中的 Bean 必须设为1
            sep: 分隔符（用于 list 类型元素分隔）
        """
        if not self.beans_file.exists():
            print(f"错误: 文件不存在 {self.beans_file}")
            return False

        # 检查是否已存在
        existing = self.get_bean(full_name)
        if existing:
            print(f"错误: Bean {full_name} 已存在")
            return False

        wb = openpyxl.load_workbook(self.beans_file)
        sheet = wb.active

        # 找到最后一行
        last_row = sheet.max_row

        # 添加 Bean 定义行
        row_num = last_row + 1
        sheet.cell(row=row_num, column=2, value=full_name)  # full_name
        sheet.cell(row=row_num, column=3, value=parent)     # parent
        sheet.cell(row=row_num, column=4, value=value_type) # valueType
        sheet.cell(row=row_num, column=5, value=sep)        # sep
        sheet.cell(row=row_num, column=6, value=alias)      # alias
        sheet.cell(row=row_num, column=7, value=comment)    # comment
        
        # 添加第一个字段
        if fields:
            field = fields[0]
            sheet.cell(row=row_num, column=10, value=field.get("name"))
            sheet.cell(row=row_num, column=11, value=field.get("alias", ""))
            sheet.cell(row=row_num, column=12, value=field.get("type"))
            sheet.cell(row=row_num, column=14, value=field.get("comment", ""))
        
        # 添加剩余字段
        for i, field in enumerate(fields[1:], start=1):
            row_num = last_row + 1 + i
            sheet.cell(row=row_num, column=10, value=field.get("name"))
            sheet.cell(row=row_num, column=11, value=field.get("alias", ""))
            sheet.cell(row=row_num, column=12, value=field.get("type"))
            sheet.cell(row=row_num, column=14, value=field.get("comment", ""))
        
        wb.save(self.beans_file)
        wb.close()
        
        print(f"✓ 已添加 Bean: {full_name}")
        return True
    
    def delete_bean(self, bean_name: str) -> bool:
        """删除 Bean"""
        if not self.beans_file.exists():
            print(f"错误: 文件不存在 {self.beans_file}")
            return False
        
        wb = openpyxl.load_workbook(self.beans_file)
        sheet = wb.active
        
        # 找到 Bean 的起始和结束行
        start_row = None
        end_row = None
        
        for i, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
            full_name = row[1]
            
            if full_name == bean_name or (full_name and full_name.endswith("." + bean_name)):
                start_row = i
            elif start_row and full_name:  # 遇到下一个 Bean
                end_row = i - 1
                break
        
        if not start_row:
            print(f"错误: 未找到 Bean {bean_name}")
            wb.close()
            return False
        
        if not end_row:
            end_row = sheet.max_row
        
        # 从后往前删除行
        for row_num in range(end_row, start_row - 1, -1):
            sheet.delete_rows(row_num)
        
        wb.save(self.beans_file)
        wb.close()
        
        print(f"✓ 已删除 Bean: {bean_name}")
        return True
    
    def update_bean(self, bean_name: str, sep: str = None, comment: str = None,
                    alias: str = None, parent: str = None, value_type: int = None) -> bool:
        """更新 Bean 属性

        Args:
            bean_name: Bean 名称
            sep: 分隔符（用于 list 类型的元素分隔）
            comment: 注释
            alias: 别名
            parent: 父类
            value_type: 是否为值类型（0=普通类，1=值类型/struct）
        """
        if not self.beans_file.exists():
            print(f"错误: 文件不存在 {self.beans_file}")
            return False
        
        wb = openpyxl.load_workbook(self.beans_file)
        sheet = wb.active
        
        # 找到 Bean 定义行
        target_row = None
        for i, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
            full_name = row[1]  # B列 = full_name
            if full_name == bean_name or (full_name and full_name.endswith("." + bean_name)):
                target_row = i
                break
        
        if not target_row:
            print(f"错误: 未找到 Bean {bean_name}")
            wb.close()
            return False
        
        # 更新属性（列索引：B=2 full_name, C=3 parent, D=4 valueType, E=5 sep, F=6 alias, G=7 comment）
        updated = []
        if sep is not None:
            sheet.cell(row=target_row, column=5, value=sep)  # E列 = sep
            updated.append(f"sep={sep}")
        if comment is not None:
            sheet.cell(row=target_row, column=7, value=comment)  # G列 = comment
            updated.append(f"comment={comment}")
        if alias is not None:
            sheet.cell(row=target_row, column=6, value=alias)  # F列 = alias
            updated.append(f"alias={alias}")
        if parent is not None:
            sheet.cell(row=target_row, column=3, value=parent)  # C列 = parent
            updated.append(f"parent={parent}")
        if value_type is not None:
            sheet.cell(row=target_row, column=4, value=value_type)  # D列 = valueType
            updated.append(f"valueType={value_type}")
        
        if not updated:
            print("警告: 没有指定要更新的属性")
            wb.close()
            return False
        
        wb.save(self.beans_file)
        wb.close()
        
        print(f"✓ 已更新 Bean: {bean_name}")
        print(f"  更新属性: {', '.join(updated)}")
        return True
    
    # ==================== 表操作 ====================
    
    def list_tables(self) -> List[Dict[str, Any]]:
        """列出所有表（包括 __tables__.xlsx 中的表和自动导入表）"""
        tables = []
        existing_names = set()

        # 1. 从 __tables__.xlsx 读取传统表定义
        if self.tables_file.exists():
            wb = openpyxl.load_workbook(self.tables_file)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=4, values_only=True):
                full_name = row[1]  # B列
                if full_name and isinstance(full_name, str):
                    tables.append({
                        "full_name": full_name,
                        "value_type": self._cell(row, 2),  # C列
                        "input": self._cell(row, 4),  # E列
                        "index": self._cell(row, 5),  # F列
                        "mode": self._cell(row, 6),  # G列
                        "comment": self._cell(row, 8),  # I列
                        "source": "__tables__.xlsx"
                    })
                    existing_names.add(full_name)

            wb.close()

        # 2. 扫描自动导入表（# 开头的文件）
        auto_tables = self._scan_auto_import_tables()
        for table in auto_tables:
            if table["full_name"] not in existing_names:
                tables.append(table)

        return tables

    def _scan_auto_import_tables(self) -> List[Dict[str, Any]]:
        """扫描自动导入表（# 开头的文件）

        命名规则：
        - #Item.xlsx -> 表名 TbItem，类型 Item
        - #Item-道具表.xlsx -> 表名 TbItem，注释 道具表
        - reward/#Reward.xlsx -> 表名 reward.TbReward
        """
        tables = []

        def scan_directory(directory: Path, module_prefix: str = ""):
            for item in directory.iterdir():
                if item.is_dir() and not item.name.startswith("."):
                    # 子目录，递归扫描
                    new_prefix = f"{module_prefix}.{item.name}" if module_prefix else item.name
                    scan_directory(item, new_prefix)
                elif item.is_file() and item.name.startswith("#") and item.suffix == ".xlsx":
                    # 跳过特殊文件
                    if item.name.startswith("##"):
                        continue

                    # 解析文件名
                    name_part = item.name[1:-5]  # 去掉 # 和 .xlsx
                    parts = name_part.split("-", 1)
                    base_name = parts[0]
                    comment = parts[1] if len(parts) > 1 else ""

                    # 构造表名
                    value_type = base_name
                    table_name = f"Tb{base_name}"
                    if module_prefix:
                        full_name = f"{module_prefix}.{table_name}"
                    else:
                        full_name = table_name

                    tables.append({
                        "full_name": full_name,
                        "value_type": value_type,
                        "input": item.relative_to(self.data_dir).as_posix(),
                        "index": "",  # 自动从表头读取
                        "mode": "",
                        "comment": comment,
                        "source": "auto-import"
                    })

        if self.data_dir.exists():
            scan_directory(self.data_dir)

        return tables
    
    def get_table(self, table_name: str) -> Optional[Dict[str, Any]]:
        """获取指定表的定义"""
        tables = self.list_tables()
        for table in tables:
            if table["full_name"] == table_name or table["full_name"].endswith("." + table_name):
                return table
        return None
    
    # 用户偏好设置
    USER_PREFERENCES = {
        "prefer_auto_import": False  # 默认使用 __tables__.xlsx 正式注册，不使用 # 自动导入
    }
    
    def set_preference(self, key: str, value: Any) -> None:
        """设置用户偏好"""
        self.USER_PREFERENCES[key] = value
        print(f"✓ 已保存偏好设置: {key} = {value}")
    
    def get_preference(self, key: str, default: Any = None) -> Any:
        """获取用户偏好"""
        return self.USER_PREFERENCES.get(key, default)
    
    def add_table(self, full_name: str, fields: List[Dict], value_type: str = "",
                  input_file: str = "", sheet_name: str = "", mode: str = "", 
                  comment: str = "", index: str = "", groups: List[str] = None,
                  auto_import: bool = None, vertical: bool = False) -> bool:
        """新增配置表
        
        Args:
            full_name: 表全名 (如 test.TbItem)
            fields: 字段列表，格式 [{"name": "id", "type": "int", "comment": "ID", "group": "c"}]
            value_type: 值类型 (如 TbItem)
            input_file: 输入文件名 (如 item.xlsx)
            sheet_name: Sheet名称（可选，默认使用表名）
            mode: 模式
            comment: 表注释
            index: 主键定义 (如 "id" 或 "id1+id2")
            groups: 分组列表 (如 ["c", "s"])
            auto_import: 是否使用自动导入格式（None 时使用用户偏好，默认 False 即在 __tables__.xlsx 注册）
            vertical: 是否使用纵表模式（适合单例表）
        """
        # 决定是否使用自动导入格式（默认在 __tables__.xlsx 正式注册）
        use_auto_import = auto_import if auto_import is not None else self.get_preference("prefer_auto_import", False)
        
        # 提取表名（去掉模块前缀）
        table_short_name = full_name.split(".")[-1] if "." in full_name else full_name
        # 去掉 Tb 前缀获取 value_type
        if table_short_name.startswith("Tb"):
            record_type = table_short_name[2:]
        else:
            record_type = table_short_name
        
        if use_auto_import:
            # 自动导入格式：#表名-注释.xlsx
            if not comment:
                comment = table_short_name  # 默认用表名作为注释
            auto_file_name = f"#{record_type}-{comment}.xlsx"
            # 处理模块前缀，创建子目录
            if "." in full_name:
                module_name = full_name.split(".")[0]
                target_dir = self.data_dir / module_name
                # 确保子目录存在
                target_dir.mkdir(parents=True, exist_ok=True)
                excel_path = target_dir / auto_file_name
            else:
                excel_path = self.data_dir / auto_file_name
            
            # 检查文件是否已存在
            if excel_path.exists():
                print(f"错误: 自动导入文件已存在 {auto_file_name}")
                return False
            
            # 创建自动导入格式的 Excel 文件
            self._create_table_excel(excel_path, fields, "Sheet1", groups, vertical)
            
            print(f"✓ 已创建自动导入表: {auto_file_name}")
            print(f"  表名: {table_short_name}")
            print(f"  记录类型: {record_type}")
            print(f"  注释: {comment}")
            if vertical:
                print(f"  模式: 纵表")
            return True
        
        # 传统方式：在 __tables__.xlsx 中注册
        # 检查是否已存在
        existing = self.get_table(full_name)
        if existing:
            print(f"错误: 表 {full_name} 已存在")
            return False
        
        # 1. 在 __tables__.xlsx 中注册表定义
        if self.tables_file.exists():
            wb = openpyxl.load_workbook(self.tables_file)
            sheet = wb.active
            last_row = sheet.max_row
            
            # 添加表定义行
            row_num = last_row + 1
            sheet.cell(row=row_num, column=2, value=full_name)  # full_name
            sheet.cell(row=row_num, column=3, value=value_type)  # value
            sheet.cell(row=row_num, column=5, value=input_file)  # input (E列)
            sheet.cell(row=row_num, column=6, value=index)  # index (F列)
            sheet.cell(row=row_num, column=7, value=mode)  # mode (G列)
            sheet.cell(row=row_num, column=9, value=comment)  # comment (I列)
            
            wb.save(self.tables_file)
            wb.close()
        
        # 2. 创建数据表 Excel 文件（或添加 sheet）
        # --input 缺省时自动推导路径
        if not input_file:
            # 从 full_name 推导: Map.TbItem → Map/item.xlsx
            # 从 record_type 推导: Item → item.xlsx
            if "." in full_name:
                module = full_name.split(".")[0]
                input_file = f"{module}/{record_type.lower()}.xlsx"
            else:
                input_file = f"{record_type.lower()}.xlsx"

        excel_path = self.data_dir / input_file
        # 确定默认 sheet 名
        if not sheet_name:
            if excel_path.exists():
                # 已有文件：默认使用表名作为 sheet 名
                sheet_name = full_name.split(".")[-1] if "." in full_name else full_name
            else:
                # 新建文件：默认使用 Sheet1（Luban 兼容）
                sheet_name = "Sheet1"
        self._create_table_excel(excel_path, fields, sheet_name, groups, vertical)

        # 3. 自动 validate 创建的 xlsx
        validate_result = self.validate_table(full_name)
        if not validate_result["valid"]:
            print(f"  ⚠ 创建后验证发现问题:")
            for err in validate_result["errors"]:
                print(f"    ✗ {err}")
            for warn in validate_result.get("warnings", []):
                print(f"    ⚠ {warn}")

        print(f"✓ 已添加表: {full_name}")
        print(f"  数据文件: {excel_path}")
        print(f"  Sheet: {sheet_name}")
        if vertical:
            print(f"  模式: 纵表")
        return True
    
    def check_legacy_tables(self) -> List[Dict[str, Any]]:
        """检查非自动导入格式的老表，返回建议转换的表列表"""
        legacy_tables = []
        
        # 检查 __tables__.xlsx 中的表
        tables = self.list_tables()
        for table in tables:
            input_file = table.get("input", "")
            if input_file and not input_file.startswith("#"):
                # 非自动导入格式
                table_name = table["full_name"]
                comment = table.get("comment", "")
                short_name = table_name.split(".")[-1]
                
                # 生成建议的自动导入文件名
                if short_name.startswith("Tb"):
                    record_type = short_name[2:]
                else:
                    record_type = short_name
                
                suggested_name = f"#{record_type}-{comment if comment else short_name}.xlsx"
                
                legacy_tables.append({
                    "current_name": table_name,
                    "current_file": input_file,
                    "suggested_file": suggested_name,
                    "comment": comment
                })
        
        return legacy_tables
    
    def suggest_auto_import_migration(self) -> None:
        """建议用户将老表迁移到自动导入格式"""
        legacy_tables = self.check_legacy_tables()
        
        if not legacy_tables:
            print("✓ 所有表都已使用自动导入格式")
            return
        
        print(f"发现 {len(legacy_tables)} 个表可以转换为自动导入格式：")
        print()
        for t in legacy_tables:
            print(f"  • {t['current_name']}")
            print(f"    当前文件: {t['current_file']}")
            print(f"    建议改为: {t['suggested_file']}")
            print()
        
        print("💡 使用以下命令转换：")
        print("   python scripts/luban_helper.py table migrate-auto --data-dir Datas")
    
    def migrate_to_auto_import(self, table_name: str = None) -> bool:
        """将表迁移到自动导入格式
        
        Args:
            table_name: 要迁移的表名（None 表示迁移所有）
        """
        legacy_tables = self.check_legacy_tables()
        
        if not legacy_tables:
            print("✓ 没有需要迁移的表")
            return True
        
        if table_name:
            # 迁移指定表
            legacy_tables = [t for t in legacy_tables if t["current_name"] == table_name or t["current_name"].endswith("." + table_name)]
            if not legacy_tables:
                print(f"错误: 表 {table_name} 不是老格式或不存在")
                return False
        
        import shutil
        
        for t in legacy_tables:
            old_path = self.data_dir / t["current_file"]
            new_path = self.data_dir / t["suggested_file"]
            
            if old_path.exists():
                # 重命名文件
                shutil.move(str(old_path), str(new_path))
                print(f"✓ 迁移: {t['current_file']} -> {t['suggested_file']}")
            else:
                print(f"⚠ 文件不存在: {t['current_file']}")
        
        # 从 __tables__.xlsx 中删除这些表的定义
        if self.tables_file.exists():
            wb = openpyxl.load_workbook(self.tables_file)
            sheet = wb.active
            
            rows_to_delete = []
            for row in range(4, sheet.max_row + 1):
                name = sheet.cell(row=row, column=2).value
                if name:
                    for t in legacy_tables:
                        if name == t["current_name"]:
                            rows_to_delete.append(row)
                            break
            
            for row in sorted(rows_to_delete, reverse=True):
                sheet.delete_rows(row)
            
            wb.save(self.tables_file)
            wb.close()
        
        print()
        print(f"✓ 已迁移 {len(legacy_tables)} 个表到自动导入格式")
        return True
    
    def _create_table_excel(self, excel_path: Path, fields: List[Dict], 
                            sheet_name: str = "Sheet1", groups: List[str] = None,
                            vertical: bool = False):
        """创建数据表 Excel 文件或添加 sheet
        
        Excel 结构:
        横表（默认）:
        行1: ##var | field1 | field2 | ...
        行2: ##type | type1 | type2 | ...
        行3+: ## | comment1 | comment2 | ... (可多行，用 | 分隔)
        行N: ##group | group1 | group2 | ... (可选)
        行N+1+: 数据行
        
        纵表:
        行1: ##column (或 ##vertical)
        行2+: 字段名 | 类型 | 注释 | 值
        """
        # 检查文件是否已存在
        if excel_path.exists():
            # 打开已有文件，添加新 sheet
            wb = openpyxl.load_workbook(excel_path)
            # 检查 sheet 是否已存在
            if sheet_name in wb.sheetnames:
                print(f"  警告: Sheet '{sheet_name}' 已存在，将覆盖")
                # 删除已有 sheet
                del wb[sheet_name]
            # 创建新 sheet
            sheet = wb.create_sheet(title=sheet_name)
        else:
            # 创建新文件
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = sheet_name
        
        if vertical:
            # 纵表模式：一行一个字段
            sheet.cell(row=1, column=1, value="##column")
            # 标题行
            sheet.cell(row=2, column=1, value="##var")
            sheet.cell(row=2, column=2, value="##type")
            sheet.cell(row=2, column=3, value="##")
            sheet.cell(row=2, column=4, value="##group")
            
            # 字段行
            for i, field in enumerate(fields, start=3):
                sheet.cell(row=i, column=1, value=field.get("name", ""))
                sheet.cell(row=i, column=2, value=field.get("type", ""))
                sheet.cell(row=i, column=3, value=field.get("comment", ""))
                sheet.cell(row=i, column=4, value=field.get("group", ""))
        else:
            # 横表模式（默认）
            # 行1: ##var 行 (字段名)
            sheet.cell(row=1, column=1, value="##var")
            for i, field in enumerate(fields, start=2):
                sheet.cell(row=1, column=i, value=field.get("name", ""))
            
            # 行2: ##type 行 (类型)
            sheet.cell(row=2, column=1, value="##type")
            for i, field in enumerate(fields, start=2):
                sheet.cell(row=2, column=i, value=field.get("type", ""))
            
            # 行3+: ## 行 (注释，支持多行，用 | 分隔)
            # 找出最大注释行数
            max_comment_rows = 1
            comment_rows = []  # 存储每个字段的多行注释列表
            for field in fields:
                comment = field.get("comment", "")
                # 用 | 分隔多行注释
                comments = [c.strip() for c in comment.split("|") if c.strip()]
                if not comments:
                    comments = [""]  # 至少有一行（空）
                comment_rows.append(comments)
                max_comment_rows = max(max_comment_rows, len(comments))
            
            # 写入多行注释
            for row_idx in range(max_comment_rows):
                row_num = 3 + row_idx
                sheet.cell(row=row_num, column=1, value="##")
                for i, comments in enumerate(comment_rows, start=2):
                    comment = comments[row_idx] if row_idx < len(comments) else ""
                    sheet.cell(row=row_num, column=i, value=comment)
            
            # ##group 行 (分组，如果字段有 group 属性则生成)
            group_row = 3 + max_comment_rows
            # 检查是否有字段有 group 属性
            has_group = any(field.get("group") for field in fields)
            if groups or has_group:
                sheet.cell(row=group_row, column=1, value="##group")
                for i, field in enumerate(fields, start=2):
                    field_group = field.get("group", "")
                    sheet.cell(row=group_row, column=i, value=field_group)
        
        # 保存文件
        wb.save(excel_path)
        wb.close()
    
    def delete_table(self, table_name: str, delete_data: bool = False) -> bool:
        """删除配置表
        
        Args:
            table_name: 表名称
            delete_data: 是否同时删除数据文件
        """
        if not self.tables_file.exists():
            print(f"错误: 文件不存在 {self.tables_file}")
            return False
        
        wb = openpyxl.load_workbook(self.tables_file)
        sheet = wb.active
        
        # 找到表定义行
        target_row = None
        target_input = None
        
        for i, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
            full_name = row[1]
            if full_name == table_name or (full_name and full_name.endswith("." + table_name)):
                target_row = i
                target_input = self._cell(row, 4, None)  # E列 = input
                break
        
        if not target_row:
            print(f"错误: 未找到表 {table_name}")
            wb.close()
            return False
        
        # 删除表定义行
        sheet.delete_rows(target_row)
        wb.save(self.tables_file)
        wb.close()
        
        # 可选：删除数据文件
        if delete_data and target_input:
            data_file = self.data_dir / str(target_input)
            if data_file.exists():
                data_file.unlink()
                print(f"✓ 已删除数据文件: {data_file}")
        
        print(f"✓ 已删除表: {table_name}")
        return True
    
    def update_table(self, table_name: str, comment: str = None, input_file: str = None,
                     mode: str = None, value_type: str = None) -> bool:
        """更新表属性
        
        Args:
            table_name: 表名称
            comment: 注释
            input_file: 输入文件名
            mode: 模式
            value_type: 值类型
        """
        if not self.tables_file.exists():
            print(f"错误: 文件不存在 {self.tables_file}")
            return False
        
        wb = openpyxl.load_workbook(self.tables_file)
        sheet = wb.active
        
        # 找到表定义行
        target_row = None
        for i, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
            full_name = row[1]  # B列
            if full_name == table_name or (full_name and full_name.endswith("." + table_name)):
                target_row = i
                break
        
        if not target_row:
            print(f"错误: 未找到表 {table_name}")
            wb.close()
            return False
        
        # 更新属性
        # 列索引：B=2 full_name, C=3 value, E=5 input, F=6 index, G=7 mode, I=9 comment
        updated = []
        if comment is not None:
            sheet.cell(row=target_row, column=9, value=comment)
            updated.append(f"comment={comment}")
        if input_file is not None:
            sheet.cell(row=target_row, column=5, value=input_file)
            updated.append(f"input={input_file}")
        if mode is not None:
            sheet.cell(row=target_row, column=7, value=mode)
            updated.append(f"mode={mode}")
        if value_type is not None:
            sheet.cell(row=target_row, column=3, value=value_type)
            updated.append(f"value_type={value_type}")
        
        if not updated:
            print("警告: 没有指定要更新的属性")
            wb.close()
            return False
        
        wb.save(self.tables_file)
        wb.close()
        
        print(f"✓ 已更新表: {table_name}")
        print(f"  更新属性: {', '.join(updated)}")
        return True
    
    # ==================== 字段操作 ====================

    def _resolve_input_path(self, input_val: str) -> Tuple[Path, Optional[str]]:
        """解析 Luban input 字段为 (excel_path, sheet_name)。

        Luban input 语法:
          - 单文件: "Map/map.xlsx" → (data_dir/Map/map.xlsx, None)
          - 多 sheet @ 语法: "战斗/技能基础@技能配置表.xlsx"
            → (data_dir/战斗/技能配置表.xlsx, "技能基础")
            即: @ 之前是 目录/sheet名，@ 之后是 文件名

        Args:
            input_val: __tables__.xlsx 中 input 列的值

        Returns:
            (Path, Optional[str]): (绝对路径, sheet名或None)
        """
        input_str = str(input_val).strip()
        if "@" in input_str:
            # Luban @ 语法: "目录/sheet名@文件名.xlsx"
            # @ 之前 = 目录/sheet名, @ 之后 = 文件名
            before_at, after_at = input_str.split("@", 1)
            # before_at 的最后一段是 sheet 名，前面是目录
            before_parts = before_at.rstrip("/").rsplit("/", 1)
            if len(before_parts) == 2:
                directory = before_parts[0]
                sheet_name = before_parts[1]
            else:
                directory = ""
                sheet_name = before_parts[0]
            # after_at 是文件名
            file_name = after_at.strip()
            if directory:
                excel_path = self.data_dir / directory / file_name
            else:
                excel_path = self.data_dir / file_name
            return (excel_path, sheet_name)
        else:
            # 无 @ 语法: 直接作为相对路径
            return (self.data_dir / input_str, None)

    def _get_table_excel_path(self, table_name: str, sheet_name: str = None) -> Optional[tuple]:
        """获取表对应的 Excel 文件路径和 sheet 名

        Returns:
            (excel_path, sheet_name) 或 None
        """
        # 先尝试从 __tables__.xlsx 查找
        table = self.get_table(table_name)

        if table:
            input_val = table.get("input", "")
            if not input_val or isinstance(input_val, bool):
                # 尝试从 mode 字段获取
                mode_str = table.get("mode", "")
                if mode_str and isinstance(mode_str, str):
                    # mode 字段直接作为文件名查找
                    resolved = self.data_dir / mode_str
                    if resolved.exists():
                        return (resolved, mode_str)
                    # 尝试 glob 查找（mode 可能只是文件名部分）
                    matches = list(self.data_dir.glob(f"**/{mode_str}"))
                    if matches:
                        return (matches[0], mode_str)
                return None

            # 使用 _resolve_input_path 直接构建路径
            excel_path, resolved_sheet = self._resolve_input_path(input_val)

            if excel_path.exists():
                # 确定 sheet 名优先级: 调用者指定 > input 解析 > None
                final_sheet = sheet_name or resolved_sheet
                return (excel_path, final_sheet)

            # 文件不存在，回退到 glob 搜索（兼容旧格式或路径拼写差异）
            file_name = excel_path.name
            matches = list(self.data_dir.glob(f"**/{file_name}"))
            if matches:
                final_sheet = sheet_name or resolved_sheet
                return (matches[0], final_sheet)

            return None

        # 尝试从自动导入表查找
        # 提取表短名
        short_name = table_name.split(".")[-1] if "." in table_name else table_name
        # 去掉 Tb 前缀
        if short_name.startswith("Tb"):
            record_type = short_name[2:]
        else:
            record_type = short_name

        # 查找 #RecordType-*.xlsx 格式的文件
        for excel_file in self.data_dir.rglob(f"#*.xlsx"):
            file_name = excel_file.stem  # 去掉 .xlsx
            # 解析文件名: #RecordType-Comment
            if file_name.startswith("#"):
                name_part = file_name[1:]
                # 提取 RecordType（去掉 -Comment 部分）
                if "-" in name_part:
                    file_record_type = name_part.split("-")[0]
                else:
                    file_record_type = name_part

                if file_record_type == record_type:
                    return (excel_file, None)

        return None
    
    def list_fields(self, table_name: str, sheet_name: str = None) -> List[Dict[str, Any]]:
        """列出表的所有字段

        数据文件缺失或解析失败时返回空列表，而非 None。
        """
        result = self._get_table_excel_path(table_name, sheet_name)
        if not result:
            return []

        excel_path, actual_sheet = result
        data = self._parse_excel_data(excel_path, actual_sheet)
        if data:
            return data.get("fields", [])
        return []
    
    def _infer_field_group(self, field_name: str, field_type: str) -> str:
        """根据字段名和类型推断默认分组
        
        Args:
            field_name: 字段名
            field_type: 字段类型
            
        Returns:
            推断的分组: 'c' (客户端), 's' (服务器), 'cs' (两者都有)
        """
        field_name_lower = field_name.lower()
        
        # 客户端专用字段关键词
        client_keywords = [
            'name', 'desc', 'description', 'icon', 'image', 'sprite', 'model',
            'effect', 'sound', 'animation', 'ui', 'display', 'show', 'hide',
            'color', 'colour', 'visual', 'appearance', 'tip', 'tips', 'tooltip',
            'title', 'text', 'label', 'prefab', 'asset', 'resource', 'path',
            'audio', 'music', 'voice', 'particle', 'shader', 'material',
            'texture', 'atlas', 'skin', 'avatar', 'portrait', 'emoji',
            'client', 'local'
        ]
        
        # 服务器专用字段关键词
        server_keywords = [
            'server', 'backend', 'logic', 'calc', 'compute', 'process',
            'internal', 'sys', 'system', 'admin', 'gm', 'debug', 'test',
            'rate', 'ratio', 'factor', 'coefficient', 'multiplier',
            'duration', 'interval', 'cooldown', 'cd', 'timeout',
            'damage', 'hp', 'mp', 'exp', 'level', 'lv', 'attack', 'defense',
            'server_id', 'channel', 'platform'
        ]
        
        # 检查是否客户端专用
        for keyword in client_keywords:
            if keyword in field_name_lower:
                return 'c'
        
        # 检查是否服务器专用
        for keyword in server_keywords:
            if keyword in field_name_lower:
                return 's'
        
        # 特殊字段名判断
        # id 通常是两边都需要
        if field_name_lower == 'id' or field_name_lower.endswith('_id'):
            return 'cs'
        
        # 类型判断
        if field_type:
            type_lower = field_type.lower()
            # 客户端相关类型
            if any(kw in type_lower for kw in ['icon', 'image', 'sprite', 'audio', 'effect', 'prefab']):
                return 'c'
        
        # 无法明确判断，默认两边都有
        return 'cs'
    
    def add_field(self, table_name: str, field_name: str, field_type: str = "",
                  field_comment: str = "", field_group: str = "",
                  sheet_name: str = None, position: int = -1) -> bool:
        """添加字段到表
        
        Args:
            table_name: 表名称
            field_name: 字段名
            field_type: 字段类型
            field_comment: 字段注释（支持多行，用 | 分隔）
            field_group: 字段分组（为空时自动推断）
            sheet_name: Sheet名（多 sheet 文件需要指定）
            position: 插入位置（-1 表示末尾）
        """
        # 自动推断分组
        if not field_group:
            field_group = self._infer_field_group(field_name, field_type)
        
        result = self._get_table_excel_path(table_name, sheet_name)
        if not result:
            print(f"错误: 未找到表 {table_name} 的数据文件")
            return False
        
        excel_path, actual_sheet = result
        
        try:
            wb = openpyxl.load_workbook(excel_path)
            
            # 确定 sheet
            if actual_sheet:
                if actual_sheet not in wb.sheetnames:
                    print(f"错误: Sheet '{actual_sheet}' 不存在")
                    wb.close()
                    return False
                sheet = wb[actual_sheet]
            else:
                sheet = wb.active
            
            # 找到当前最大列
            max_col = sheet.max_column
            
            # 检查字段是否已存在
            for col in range(2, max_col + 1):
                if sheet.cell(row=1, column=col).value == field_name:
                    print(f"错误: 字段 '{field_name}' 已存在")
                    wb.close()
                    return False
            
            # 确定插入位置
            if position < 0 or position >= max_col:
                insert_col = max_col + 1
            else:
                insert_col = position + 2  # +2 因为 A列是 ##var
            
            # 插入列
            if insert_col <= max_col:
                sheet.insert_cols(insert_col)
            
            # 设置字段值
            sheet.cell(row=1, column=insert_col, value=field_name)  # ##var
            sheet.cell(row=2, column=insert_col, value=field_type)  # ##type
            
            # 找出注释行数和 ##group 行位置
            comment_rows = []
            group_row = None
            for row in range(3, sheet.max_row + 1):
                cell_value = sheet.cell(row=row, column=1).value
                if cell_value == "##":
                    comment_rows.append(row)
                elif cell_value == "##group":
                    group_row = row
                    break
                elif cell_value and not cell_value.startswith("##"):
                    break  # 数据行开始
            
            # 解析多行注释
            comments = [c.strip() for c in field_comment.split("|") if c.strip()]
            if not comments:
                comments = [""]
            
            # 如果新字段注释行数大于现有注释行数，需要插入新行
            if len(comments) > len(comment_rows):
                # 需要插入额外的注释行
                extra_rows = len(comments) - len(comment_rows)
                # 在最后一个注释行之后插入
                insert_after = comment_rows[-1] if comment_rows else 2
                for _ in range(extra_rows):
                    sheet.insert_rows(insert_after + 1)
                    # 更新后续行号
                    if group_row:
                        group_row += 1
                # 更新 comment_rows
                for i in range(len(comment_rows), len(comments)):
                    comment_rows.append(insert_after + 1 + i - len(comment_rows))
                    sheet.cell(row=comment_rows[-1], column=1, value="##")
            
            # 写入注释
            for i, comment in enumerate(comments):
                if i < len(comment_rows):
                    sheet.cell(row=comment_rows[i], column=insert_col, value=comment)
                else:
                    # 超出现有注释行数，写入最后一行注释
                    sheet.cell(row=comment_rows[-1], column=insert_col, value=comment)
            
            # 填充剩余注释行为空（如果新字段注释行数少于现有注释行数）
            for i in range(len(comments), len(comment_rows)):
                sheet.cell(row=comment_rows[i], column=insert_col, value="")
            
            # 设置分组
            if group_row:
                sheet.cell(row=group_row, column=insert_col, value=field_group)
            
            wb.save(excel_path)
            wb.close()
            
            print(f"✓ 已添加字段: {field_name} ({field_type})")
            return True
            
        except Exception as e:
            print(f"错误: {e}")
            return False
    
    def update_field(self, table_name: str, field_name: str,
                     new_name: str = None, new_type: str = None,
                     new_comment: str = None, new_group: str = None,
                     sheet_name: str = None) -> bool:
        """修改字段
        
        Args:
            table_name: 表名称
            field_name: 原字段名
            new_name: 新字段名（可选）
            new_type: 新类型（可选）
            new_comment: 新注释（可选，支持多行，用 | 分隔）
            new_group: 新分组（可选）
            sheet_name: Sheet名（多 sheet 文件需要指定）
        """
        result = self._get_table_excel_path(table_name, sheet_name)
        if not result:
            print(f"错误: 未找到表 {table_name} 的数据文件")
            return False
        
        excel_path, actual_sheet = result
        
        try:
            wb = openpyxl.load_workbook(excel_path)
            
            if actual_sheet:
                if actual_sheet not in wb.sheetnames:
                    print(f"错误: Sheet '{actual_sheet}' 不存在")
                    wb.close()
                    return False
                sheet = wb[actual_sheet]
            else:
                sheet = wb.active
            
            # 找到字段列
            field_col = None
            for col in range(2, sheet.max_column + 1):
                if sheet.cell(row=1, column=col).value == field_name:
                    field_col = col
                    break
            
            if not field_col:
                print(f"错误: 未找到字段 '{field_name}'")
                wb.close()
                return False
            
            # 修改字段名和类型
            if new_name:
                sheet.cell(row=1, column=field_col, value=new_name)
            if new_type is not None:
                sheet.cell(row=2, column=field_col, value=new_type)
            
            # 修改注释（支持多行）
            if new_comment is not None:
                # 找出注释行
                comment_rows = []
                group_row = None
                for row in range(3, sheet.max_row + 1):
                    cell_value = sheet.cell(row=row, column=1).value
                    if cell_value == "##":
                        comment_rows.append(row)
                    elif cell_value == "##group":
                        group_row = row
                        break
                    elif cell_value and not cell_value.startswith("##"):
                        break
                
                # 解析多行注释
                comments = [c.strip() for c in new_comment.split("|") if c.strip()]
                if not comments:
                    comments = [""]
                
                # 如果新注释行数大于现有注释行数，需要插入新行
                if len(comments) > len(comment_rows):
                    extra_rows = len(comments) - len(comment_rows)
                    insert_after = comment_rows[-1] if comment_rows else 2
                    for _ in range(extra_rows):
                        sheet.insert_rows(insert_after + 1)
                        if group_row:
                            group_row += 1
                    for i in range(len(comment_rows), len(comments)):
                        comment_rows.append(insert_after + 1 + i - len(comment_rows))
                        sheet.cell(row=comment_rows[-1], column=1, value="##")
                
                # 写入注释
                for i, comment in enumerate(comments):
                    if i < len(comment_rows):
                        sheet.cell(row=comment_rows[i], column=field_col, value=comment)
                
                # 填充剩余注释行为空
                for i in range(len(comments), len(comment_rows)):
                    sheet.cell(row=comment_rows[i], column=field_col, value="")
            
            # 修改分组
            if new_group is not None:
                # 找 ##group 行
                group_row = None
                for row in range(3, sheet.max_row + 1):
                    if sheet.cell(row=row, column=1).value == "##group":
                        group_row = row
                        break
                if group_row:
                    sheet.cell(row=group_row, column=field_col, value=new_group)
            
            wb.save(excel_path)
            wb.close()
            
            print(f"✓ 已修改字段: {field_name}")
            return True
            
        except Exception as e:
            print(f"错误: {e}")
            return False
    
    def delete_field(self, table_name: str, field_name: str,
                     sheet_name: str = None, force: bool = False) -> bool:
        """删除字段（危险操作）
        
        Args:
            table_name: 表名称
            field_name: 字段名
            sheet_name: Sheet名（多 sheet 文件需要指定）
            force: 强制删除，跳过确认
        """
        result = self._get_table_excel_path(table_name, sheet_name)
        if not result:
            print(f"错误: 未找到表 {table_name} 的数据文件")
            return False
        
        excel_path, actual_sheet = result
        
        try:
            wb = openpyxl.load_workbook(excel_path)
            
            if actual_sheet:
                if actual_sheet not in wb.sheetnames:
                    print(f"错误: Sheet '{actual_sheet}' 不存在")
                    wb.close()
                    return False
                sheet = wb[actual_sheet]
            else:
                sheet = wb.active
            
            # 找到字段列
            field_col = None
            for col in range(2, sheet.max_column + 1):
                if sheet.cell(row=1, column=col).value == field_name:
                    field_col = col
                    break
            
            if not field_col:
                print(f"错误: 未找到字段 '{field_name}'")
                wb.close()
                return False
            
            # 检查是否有数据
            has_data = False
            for row in range(5, sheet.max_row + 1):  # 从第5行开始是数据
                if sheet.cell(row=row, column=field_col).value is not None:
                    has_data = True
                    break
            
            # 危险操作确认
            if not force:
                if has_data:
                    print(f"⚠️  警告: 字段 '{field_name}' 包含数据，删除后数据将丢失！")
                confirm = input(f"确认删除字段 '{field_name}'? (yes/N): ")
                if confirm.lower() != "yes":
                    print("操作已取消")
                    wb.close()
                    return False
            
            # 删除列
            sheet.delete_cols(field_col)
            
            wb.save(excel_path)
            wb.close()
            
            print(f"✓ 已删除字段: {field_name}")
            return True
            
        except Exception as e:
            print(f"错误: {e}")
            return False
    
    def disable_field(self, table_name: str, field_name: str, sheet_name: str = None) -> bool:
        """禁用字段（注释列，不导出但保留数据）
        
        通过在字段名前添加 ## 前缀来实现，Luban 会忽略以 ## 开头的字段
        
        Args:
            table_name: 表名称
            field_name: 字段名
            sheet_name: Sheet名（多 sheet 文件需要指定）
        """
        result = self._get_table_excel_path(table_name, sheet_name)
        if not result:
            print(f"错误: 未找到表 {table_name} 的数据文件")
            return False
        
        excel_path, actual_sheet = result
        
        try:
            wb = openpyxl.load_workbook(excel_path)
            
            if actual_sheet:
                if actual_sheet not in wb.sheetnames:
                    print(f"错误: Sheet '{actual_sheet}' 不存在")
                    wb.close()
                    return False
                sheet = wb[actual_sheet]
            else:
                sheet = wb.active
            
            # 找到字段列
            field_col = None
            for col in range(2, sheet.max_column + 1):
                cell_value = sheet.cell(row=1, column=col).value
                if cell_value == field_name:
                    field_col = col
                    break
            
            if not field_col:
                print(f"错误: 未找到字段 '{field_name}'")
                wb.close()
                return False
            
            # 检查是否已被禁用
            current_name = sheet.cell(row=1, column=field_col).value
            if current_name and current_name.startswith("##"):
                print(f"字段 '{field_name}' 已被禁用")
                wb.close()
                return True
            
            # 在字段名前添加 ## 前缀
            sheet.cell(row=1, column=field_col, value=f"##{field_name}")
            
            wb.save(excel_path)
            wb.close()
            
            print(f"✓ 已禁用字段: {field_name} (导表时将被忽略)")
            return True
            
        except Exception as e:
            print(f"错误: {e}")
            return False
    
    def enable_field(self, table_name: str, field_name: str, sheet_name: str = None) -> bool:
        """启用字段（取消注释列）
        
        移除字段名前的 ## 前缀
        
        Args:
            table_name: 表名称
            field_name: 字段名（可以带 ## 前缀或不带）
            sheet_name: Sheet名（多 sheet 文件需要指定）
        """
        result = self._get_table_excel_path(table_name, sheet_name)
        if not result:
            print(f"错误: 未找到表 {table_name} 的数据文件")
            return False
        
        excel_path, actual_sheet = result
        
        try:
            wb = openpyxl.load_workbook(excel_path)
            
            if actual_sheet:
                if actual_sheet not in wb.sheetnames:
                    print(f"错误: Sheet '{actual_sheet}' 不存在")
                    wb.close()
                    return False
                sheet = wb[actual_sheet]
            else:
                sheet = wb.active
            
            # 找到字段列（支持带 ## 前缀或不带）
            field_col = None
            actual_field_name = None
            for col in range(2, sheet.max_column + 1):
                cell_value = sheet.cell(row=1, column=col).value
                if cell_value == field_name or cell_value == f"##{field_name}":
                    field_col = col
                    actual_field_name = cell_value
                    break
            
            if not field_col:
                print(f"错误: 未找到字段 '{field_name}'")
                wb.close()
                return False
            
            # 检查是否已被禁用
            if not actual_field_name.startswith("##"):
                print(f"字段 '{field_name}' 未被禁用")
                wb.close()
                return True
            
            # 移除 ## 前缀
            new_name = actual_field_name[2:]  # 去掉前两个字符 ##
            sheet.cell(row=1, column=field_col, value=new_name)
            
            wb.save(excel_path)
            wb.close()
            
            print(f"✓ 已启用字段: {new_name}")
            return True
            
        except Exception as e:
            print(f"错误: {e}")
            return False
    
    # ==================== 数据行操作 ====================
    
    def _parse_luban_sheet(self, sheet) -> LubanSheetStructure:
        """统一解析 Luban xlsx 工作表结构

        扫描标题行（##var/##type/##group/##），提取字段定义，
        检测结构错误（缺失行、列不对齐、数据行 A 列非空等）。

        Args:
            sheet: openpyxl worksheet 对象

        Returns:
            LubanSheetStructure 标准化解析结果
        """
        result = LubanSheetStructure()

        # Phase 1: 扫描标题行
        var_row = None
        type_row = None
        group_row = None
        comment_rows = []
        last_header_row = 0

        for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
            cell_a = row[0] if row else None
            if cell_a and str(cell_a).startswith("##"):
                if cell_a == "##var":
                    if var_row is not None:
                        result.format_warnings.append(f"第 {i} 行: 重复的 ##var 行（首次出现在第 {result.var_row_num} 行）")
                    result.var_row_num = i
                    var_row = list(row)
                elif cell_a == "##type":
                    result.type_row_num = i
                    type_row = list(row)
                elif cell_a == "##group":
                    result.group_row_num = i
                    group_row = list(row)
                elif cell_a == "##":
                    comment_rows.append(list(row))
                    result.comment_row_nums.append(i)
                else:
                    # 未知 ## 标记（如 ##column 纵表）
                    result.format_warnings.append(f"第 {i} 行: 未知标题标记 '{cell_a}'")
                last_header_row = i
            else:
                # 第一个非 ## 行 = 数据起始
                result.data_start_row = i
                break

        if result.data_start_row == 0:
            result.data_start_row = last_header_row + 1

        # Phase 2: 结构完整性检查
        if result.var_row_num == 0:
            result.format_errors.append("缺少 ##var 行（Luban 必需）")
        if result.type_row_num is None:
            result.format_errors.append("缺少 ##type 行（Luban 必需，字段类型定义）")
        if result.group_row_num is None:
            result.format_warnings.append("缺少 ##group 行（建议添加，Luban 分组策略可能需要）")
        if not comment_rows:
            result.format_warnings.append("缺少 ## 注释行（建议添加，提升可读性）")

        # Phase 3: 列对齐检查
        if var_row and type_row:
            var_count = len([v for v in var_row[1:] if v])
            type_count = len(type_row) - 1  # 减去 ##type 标记列
            if var_count > type_count:
                result.format_errors.append(
                    f"##var 有 {var_count} 个字段，##type 只有 {type_count} 列，"
                    f"缺少 {var_count - type_count} 个字段类型定义"
                )

        # Phase 4: 提取字段定义
        if var_row and type_row:
            for j in range(1, len(var_row)):
                field_name = var_row[j] if j < len(var_row) else None
                field_type = type_row[j] if j < len(type_row) else None

                if field_name:
                    # 检查字段类型是否为空
                    if not field_type:
                        result.format_errors.append(f"字段 '{field_name}' 缺少类型定义（##type 行第 {j+1} 列为空）")

                    # 合并注释行
                    comments = []
                    for desc_row in comment_rows:
                        if j < len(desc_row) and desc_row[j]:
                            comments.append(str(desc_row[j]))

                    # 提取分组
                    field_group = None
                    if group_row and j < len(group_row):
                        field_group = group_row[j]

                    result.fields.append({
                        "index": j - 1,
                        "name": field_name,
                        "type": field_type or "",
                        "comment": " ".join(comments) if comments else "",
                        "group": field_group
                    })

        # Phase 5: 数据行 A 列检查
        if result.data_start_row > 0:
            for row in sheet.iter_rows(min_row=result.data_start_row, values_only=True):
                if all(c is None for c in row):
                    continue  # 跳过空行
                if row[0] is not None:
                    # A 列有值 — 可能是标签行(#tag)或格式错误
                    val_a = str(row[0])
                    if val_a.startswith("#") and not val_a.startswith("##"):
                        # 数据标签行，合法
                        pass
                    else:
                        result.format_errors.append(
                            f"数据行 A 列应为空，实际值为 '{val_a}'"
                            f"（Luban 要求字段值从 B 列开始）"
                        )
                        break  # 只报第一个

        # 保存原始行数据供后续使用
        result.var_row = var_row
        result.type_row = type_row
        result.group_row = group_row

        return result

    def _get_data_start_row(self, sheet) -> int:
        """获取数据起始行号"""
        last_header_row = 4  # 默认头部4行（##var, ##type, ##, ##group）

        for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
            cell_value = row[0]
            if cell_value and str(cell_value).startswith("##"):
                last_header_row = i
            elif cell_value is None or cell_value == "":
                # 检查该行其他列是否有数据
                if any(c is not None and c != "" for c in row[1:]):
                    # 找到数据行（A列为空，但其他列有数据）
                    return i
                # 真正的空行，继续查找
                continue
            else:
                # 找到第一个非 ## 开头且非空的行
                return i

        # 如果没找到，返回头部行+1
        return last_header_row + 1
    
    def list_rows(self, table_name: str, sheet_name: str = None, 
                  start: int = 0, limit: int = 100) -> Optional[List[Dict[str, Any]]]:
        """列出表的数据行
        
        Args:
            table_name: 表名称
            sheet_name: Sheet名
            start: 起始行索引（从0开始）
            limit: 返回行数限制
        """
        result = self._get_table_excel_path(table_name, sheet_name)
        if not result:
            print(f"错误: 未找到表 {table_name} 的数据文件")
            return None
        
        excel_path, actual_sheet = result
        
        try:
            wb = openpyxl.load_workbook(excel_path)
            
            if actual_sheet:
                if actual_sheet not in wb.sheetnames:
                    print(f"错误: Sheet '{actual_sheet}' 不存在")
                    wb.close()
                    return None
                sheet = wb[actual_sheet]
            else:
                sheet = wb.active
            
            # 获取字段名
            var_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            fields = [f for f in var_row[1:] if f and not str(f).startswith("##")]
            
            # 获取数据起始行
            data_start = self._get_data_start_row(sheet)
            
            # 读取数据行
            rows = []
            row_idx = 0
            for row in sheet.iter_rows(min_row=data_start, values_only=True):
                # 跳过空行
                if all(c is None for c in row[1:len(fields)+1]):
                    continue
                
                row_data = {"_index": row_idx}
                for i, field in enumerate(fields):
                    if i + 1 < len(row):
                        row_data[field] = row[i + 1]
                
                if row_idx >= start:
                    rows.append(row_data)
                
                row_idx += 1
                if len(rows) >= limit:
                    break
            
            wb.close()
            return rows
            
        except Exception as e:
            print(f"错误: {e}")
            return None
    
    def get_row(self, table_name: str, field: str, value: Any, 
                sheet_name: str = None) -> Optional[Dict[str, Any]]:
        """按字段值查询数据行
        
        Args:
            table_name: 表名称
            field: 字段名
            value: 字段值
            sheet_name: Sheet名
        """
        result = self._get_table_excel_path(table_name, sheet_name)
        if not result:
            print(f"错误: 未找到表 {table_name} 的数据文件")
            return None
        
        excel_path, actual_sheet = result
        
        try:
            wb = openpyxl.load_workbook(excel_path)
            
            if actual_sheet:
                if actual_sheet not in wb.sheetnames:
                    print(f"错误: Sheet '{actual_sheet}' 不存在")
                    wb.close()
                    return None
                sheet = wb[actual_sheet]
            else:
                sheet = wb.active
            
            # 检查是否是纵表（A1 为 ##column 或 ##vertical）
            a1_value = sheet.cell(row=1, column=1).value
            is_vertical = a1_value in ("##column", "##vertical")
            
            if is_vertical:
                # 纵表模式：按字段名查询值
                return self._get_vertical_table_value(sheet, field, wb)
            
            # 横表模式（默认）
            # 获取字段名
            var_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            all_fields = list(var_row[1:])
            fields = [f for f in all_fields if f and not str(f).startswith("##")]
            
            # 查找目标字段索引
            field_idx = None
            for i, f in enumerate(fields):
                if f == field:
                    field_idx = i
                    break
            
            if field_idx is None:
                print(f"错误: 字段 '{field}' 不存在")
                print(f"可用字段: {', '.join(fields)}")
                wb.close()
                return None
            
            # 获取数据起始行
            data_start = self._get_data_start_row(sheet)
            
            # 查找匹配行
            for row in sheet.iter_rows(min_row=data_start, values_only=True):
                # 跳过空行
                if all(c is None for c in row[1:len(fields)+1]):
                    continue
                
                # 比较字段值
                cell_value = row[field_idx + 1] if field_idx + 1 < len(row) else None
                
                # 类型转换比较
                if self._compare_value(cell_value, value):
                    row_data = {}
                    for i, f in enumerate(fields):
                        if i + 1 < len(row):
                            row_data[f] = row[i + 1]
                    wb.close()
                    return row_data
            
            wb.close()
            print(f"未找到 {field}={value} 的记录")
            return None
            
        except Exception as e:
            print(f"错误: {e}")
            return None
    
    def _get_vertical_table_value(self, sheet, field_name: str, wb) -> Optional[Dict[str, Any]]:
        """从纵表获取字段值
        
        纵表结构:
        行1: ##column
        行2: ##var | ##type | ## | ##group
        行3+: 字段名 | 类型 | 注释 | 分组 | 值
        """
        # 从第3行开始查找字段名
        for row in sheet.iter_rows(min_row=3, values_only=True):
            if row[0] == field_name:
                result = {
                    "name": row[0],
                    "type": self._cell(row, 1, None),
                    "comment": self._cell(row, 2, None),
                    "group": self._cell(row, 3, None),
                    "value": self._cell(row, 4, None)
                }
                wb.close()
                return result
        
        # 列出可用字段
        available_fields = []
        for row in sheet.iter_rows(min_row=3, values_only=True):
            if row[0]:
                available_fields.append(row[0])
        
        wb.close()
        print(f"错误: 字段 '{field_name}' 不存在")
        print(f"可用字段: {', '.join(available_fields)}")
        return None
    
    def _compare_value(self, cell_value: Any, search_value: Any) -> bool:
        """比较单元格值和搜索值"""
        if cell_value is None:
            return False
        
        # 直接比较
        if cell_value == search_value:
            return True
        
        # 字符串比较
        str_cell = str(cell_value).strip()
        str_search = str(search_value).strip()
        if str_cell == str_search:
            return True
        
        # 数字比较
        try:
            if float(cell_value) == float(search_value):
                return True
        except (ValueError, TypeError):
            pass
        
        return False
    
    def query_rows(self, table_name: str, conditions: Dict[str, Any],
                   sheet_name: str = None, limit: int = 100) -> Optional[List[Dict[str, Any]]]:
        """按多条件查询数据行
        
        Args:
            table_name: 表名称
            conditions: 查询条件字典 {字段名: 值}
            sheet_name: Sheet名
            limit: 返回行数限制
        """
        result = self._get_table_excel_path(table_name, sheet_name)
        if not result:
            print(f"错误: 未找到表 {table_name} 的数据文件")
            return None
        
        excel_path, actual_sheet = result
        
        try:
            wb = openpyxl.load_workbook(excel_path)
            
            if actual_sheet:
                if actual_sheet not in wb.sheetnames:
                    print(f"错误: Sheet '{actual_sheet}' 不存在")
                    wb.close()
                    return None
                sheet = wb[actual_sheet]
            else:
                sheet = wb.active
            
            # 获取字段名
            var_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            all_fields = list(var_row[1:])
            fields = [f for f in all_fields if f and not str(f).startswith("##")]
            
            # 查找条件字段索引
            cond_indices = {}
            for cond_field in conditions.keys():
                field_idx = None
                for i, f in enumerate(fields):
                    if f == cond_field:
                        field_idx = i
                        break
                if field_idx is None:
                    print(f"错误: 字段 '{cond_field}' 不存在")
                    wb.close()
                    return None
                cond_indices[cond_field] = field_idx
            
            # 获取数据起始行
            data_start = self._get_data_start_row(sheet)
            
            # 查找匹配行
            rows = []
            for row in sheet.iter_rows(min_row=data_start, values_only=True):
                # 跳过空行
                if all(c is None for c in row[1:len(fields)+1]):
                    continue
                
                # 检查所有条件
                match = True
                for cond_field, cond_value in conditions.items():
                    field_idx = cond_indices[cond_field]
                    cell_value = row[field_idx + 1] if field_idx + 1 < len(row) else None
                    if not self._compare_value(cell_value, cond_value):
                        match = False
                        break
                
                if match:
                    row_data = {}
                    for i, f in enumerate(fields):
                        if i + 1 < len(row):
                            row_data[f] = row[i + 1]
                    rows.append(row_data)
                    if len(rows) >= limit:
                        break
            
            wb.close()
            return rows
            
        except Exception as e:
            print(f"错误: {e}")
            return None
    
    def add_row(self, table_name: str, data: Dict[str, Any], 
                sheet_name: str = None, smart_insert: bool = True) -> bool:
        """添加数据行（支持智能插入）
        
        Args:
            table_name: 表名称
            data: 行数据字典
            sheet_name: Sheet名
            smart_insert: 是否启用智能插入（按ID排序）
        """
        result = self._get_table_excel_path(table_name, sheet_name)
        if not result:
            print(f"错误: 未找到表 {table_name} 的数据文件")
            return False
        
        excel_path, actual_sheet = result
        
        try:
            wb = openpyxl.load_workbook(excel_path)
            
            if actual_sheet:
                if actual_sheet not in wb.sheetnames:
                    print(f"错误: Sheet '{actual_sheet}' 不存在")
                    wb.close()
                    return False
                sheet = wb[actual_sheet]
            else:
                sheet = wb.active
            
            # 获取字段名
            var_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            fields = list(var_row[1:])  # 跳过第一列 ##var
            
            # 找到数据起始行
            data_start = self._get_data_start_row(sheet)
            
            # 智能插入：按 ID 排序
            insert_row = None
            primary_key = "id"  # 默认主键字段
            
            if smart_insert and primary_key in data:
                try:
                    new_id = int(data[primary_key]) if data[primary_key] is not None else None
                    if new_id is not None:
                        # 收集现有数据的 ID 和行号
                        existing_ids = []
                        id_col_idx = None
                        
                        # 找到 ID 列的索引
                        for i, field in enumerate(fields):
                            if field == primary_key:
                                id_col_idx = i + 2  # 列号（从1开始，跳过第一列）
                                break
                        
                        if id_col_idx:
                            # 遍历现有数据，收集 ID
                            max_id = None
                            for row_num in range(data_start, sheet.max_row + 1):
                                # 跳过标记为 ## 的注释行
                                first_col = sheet.cell(row=row_num, column=1).value
                                if first_col == "##":
                                    continue
                                cell_val = sheet.cell(row=row_num, column=id_col_idx).value
                                if cell_val is not None:
                                    try:
                                        existing_id = int(cell_val)
                                        existing_ids.append((existing_id, row_num))
                                        if max_id is None or existing_id > max_id:
                                            max_id = existing_id
                                    except (ValueError, TypeError):
                                        pass
                            
                            # 判断插入位置
                            if max_id is None or new_id > max_id:
                                # ID 最大，追加到末尾（但在 ## 注释行之前）
                                # 找到最后一个非 ## 注释行的位置
                                insert_row = sheet.max_row + 1
                                for r in range(sheet.max_row, data_start - 1, -1):
                                    first_col_val = sheet.cell(row=r, column=1).value
                                    if first_col_val == "##":
                                        insert_row = r
                                    elif any(sheet.cell(row=r, column=c).value is not None for c in range(2, sheet.max_column + 1)):
                                        # 找到有数据的行，在其后插入
                                        insert_row = r + 1
                                        break
                            else:
                                # 找到合适的插入位置
                                for i, (existing_id, row_num) in enumerate(existing_ids):
                                    if new_id < existing_id:
                                        # 插入到这一行之前
                                        insert_row = row_num
                                        # 插入新行
                                        sheet.insert_rows(insert_row)
                                        break
                                
                                if insert_row is None:
                                    # 插入到末尾
                                    insert_row = sheet.max_row + 1
                except (ValueError, TypeError):
                    pass
            
            # 如果没有智能插入，使用追加模式
            if insert_row is None:
                insert_row = sheet.max_row + 1
                # 找到实际最后一行有数据的位置
                for row in sheet.iter_rows(min_row=data_start, values_only=True):
                    if any(c is not None for c in row[1:]):
                        pass  # 继续查找
                # 追加到末尾
            
            # 写入数据
            for i, field in enumerate(fields):
                if field and not str(field).startswith("##"):
                    value = data.get(field)
                    if value is not None:
                        sheet.cell(row=insert_row, column=i + 2, value=value)
            
            wb.save(excel_path)
            wb.close()
            
            if smart_insert and primary_key in data:
                print(f"✓ 已添加数据行（ID: {data.get(primary_key)}）")
            else:
                print(f"✓ 已添加数据行")
            return True
            
        except Exception as e:
            print(f"错误: {e}")
            return False
    
    def update_row(self, table_name: str, row_index: int, data: Dict[str, Any],
                   sheet_name: str = None) -> bool:
        """更新数据行
        
        Args:
            table_name: 表名称
            row_index: 行索引（从0开始）
            data: 更新数据字典
            sheet_name: Sheet名
        """
        result = self._get_table_excel_path(table_name, sheet_name)
        if not result:
            print(f"错误: 未找到表 {table_name} 的数据文件")
            return False
        
        excel_path, actual_sheet = result
        
        try:
            wb = openpyxl.load_workbook(excel_path)
            
            if actual_sheet:
                if actual_sheet not in wb.sheetnames:
                    print(f"错误: Sheet '{actual_sheet}' 不存在")
                    wb.close()
                    return False
                sheet = wb[actual_sheet]
            else:
                sheet = wb.active
            
            # 获取字段名
            var_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            fields = list(var_row[1:])
            
            # 找到目标行
            data_start = self._get_data_start_row(sheet)
            target_row = data_start + row_index
            
            if target_row > sheet.max_row:
                print(f"错误: 行索引 {row_index} 超出范围")
                wb.close()
                return False
            
            # 更新数据
            for i, field in enumerate(fields):
                if field and not str(field).startswith("##") and field in data:
                    sheet.cell(row=target_row, column=i + 2, value=data[field])
            
            wb.save(excel_path)
            wb.close()
            
            print(f"✓ 已更新行 {row_index}")
            return True
            
        except Exception as e:
            print(f"错误: {e}")
            return False
    
    def delete_row(self, table_name: str, row_index: int, 
                   sheet_name: str = None, force: bool = False) -> bool:
        """删除数据行
        
        Args:
            table_name: 表名称
            row_index: 行索引（从0开始）
            sheet_name: Sheet名
            force: 强制删除，跳过确认
        """
        result = self._get_table_excel_path(table_name, sheet_name)
        if not result:
            print(f"错误: 未找到表 {table_name} 的数据文件")
            return False
        
        excel_path, actual_sheet = result
        
        try:
            wb = openpyxl.load_workbook(excel_path)
            
            if actual_sheet:
                if actual_sheet not in wb.sheetnames:
                    print(f"错误: Sheet '{actual_sheet}' 不存在")
                    wb.close()
                    return False
                sheet = wb[actual_sheet]
            else:
                sheet = wb.active
            
            # 找到目标行
            data_start = self._get_data_start_row(sheet)
            target_row = data_start + row_index
            
            if target_row > sheet.max_row:
                print(f"错误: 行索引 {row_index} 超出范围")
                wb.close()
                return False
            
            # 确认删除
            if not force:
                confirm = input(f"确认删除行 {row_index}？(y/n): ")
                if confirm.lower() != 'y':
                    print("已取消")
                    wb.close()
                    return False
            
            # 删除行
            sheet.delete_rows(target_row)
            
            wb.save(excel_path)
            wb.close()
            
            print(f"✓ 已删除行 {row_index}")
            return True
            
        except Exception as e:
            print(f"错误: {e}")
            return False
    
    # ==================== 批量操作 ====================
    
    def batch_add_fields(self, table_name: str, fields: List[Dict[str, Any]], 
                         sheet_name: str = None) -> bool:
        """批量添加字段
        
        Args:
            table_name: 表名称
            fields: 字段列表，格式 [{"name": "field1", "type": "int", "comment": "注释"}, ...]
            sheet_name: Sheet名
        """
        result = self._get_table_excel_path(table_name, sheet_name)
        if not result:
            print(f"错误: 未找到表 {table_name} 的数据文件")
            return False
        
        success_count = 0
        for field in fields:
            field_name = field.get("name", "")
            field_type = field.get("type", "")
            field_comment = field.get("comment", "")
            field_group = field.get("group", "")
            
            if self.add_field(table_name, field_name, field_type, field_comment, 
                             field_group, sheet_name):
                success_count += 1
        
        print(f"✓ 批量添加完成: {success_count}/{len(fields)} 个字段")
        return success_count > 0
    
    def batch_add_rows(self, table_name: str, rows: List[Dict[str, Any]],
                       sheet_name: str = None) -> bool:
        """批量添加数据行
        
        Args:
            table_name: 表名称
            rows: 数据行列表
            sheet_name: Sheet名
        """
        result = self._get_table_excel_path(table_name, sheet_name)
        if not result:
            print(f"错误: 未找到表 {table_name} 的数据文件")
            return False
        
        success_count = 0
        for data in rows:
            if self.add_row(table_name, data, sheet_name):
                success_count += 1
        
        print(f"✓ 批量添加完成: {success_count}/{len(rows)} 行数据")
        return success_count > 0
    
    # ==================== 导入导出 ====================
    
    def export_json(self, table_name: str, output_file: str = None, 
                    sheet_name: str = None) -> bool:
        """导出表数据为 JSON
        
        Args:
            table_name: 表名称
            output_file: 输出文件路径（默认打印到控制台）
            sheet_name: Sheet名
        """
        rows = self.list_rows(table_name, sheet_name, limit=10000)
        if rows is None:
            return False
        
        # 获取字段信息
        fields_info = self.list_fields(table_name, sheet_name)
        
        export_data = {
            "table": table_name,
            "fields": fields_info,
            "rows": rows
        }
        
        json_str = json.dumps(export_data, ensure_ascii=False, indent=2)
        
        if output_file:
            with open(output_file, 'w', encoding='utf-8') as f:
                f.write(json_str)
            print(f"✓ 已导出到: {output_file}")
        else:
            print(json_str)
        
        return True
    
    def import_json(self, table_name: str, input_file: str, 
                    sheet_name: str = None, mode: str = "append") -> bool:
        """从 JSON 导入数据
        
        Args:
            table_name: 表名称
            input_file: 输入文件路径
            sheet_name: Sheet名
            mode: 导入模式 - append(追加) 或 replace(替换)
        """
        try:
            with open(input_file, 'r', encoding='utf-8') as f:
                import_data = json.load(f)
        except Exception as e:
            print(f"错误: 读取文件失败 - {e}")
            return False
        
        rows = import_data.get("rows", [])
        if not rows:
            print("警告: 没有数据需要导入")
            return False
        
        # TODO: 实现 replace 模式（需要先清空表）
        if mode == "replace":
            print("警告: replace 模式暂未实现，使用 append 模式")
        
        success_count = 0
        for row in rows:
            # 移除 _index 字段
            data = {k: v for k, v in row.items() if not k.startswith("_")}
            if self.add_row(table_name, data, sheet_name):
                success_count += 1
        
        print(f"✓ 导入完成: {success_count}/{len(rows)} 行")
        return success_count > 0
    
    # ==================== 验证功能 ====================
    
    def validate_table(self, table_name: str, sheet_name: str = None) -> Dict[str, Any]:
        """验证表数据（基于 _parse_luban_sheet 统一解析器）

        Args:
            table_name: 表名称
            sheet_name: Sheet名

        Returns:
            验证结果 {"valid": bool, "errors": [], "warnings": []}
        """
        result = {
            "valid": True,
            "errors": [],
            "warnings": []
        }

        # 检查表是否存在
        result_path = self._get_table_excel_path(table_name, sheet_name)
        if not result_path:
            result["valid"] = False
            result["errors"].append(f"表 {table_name} 不存在或没有对应的数据文件")
            return result

        excel_path, actual_sheet = result_path

        try:
            wb = openpyxl.load_workbook(excel_path)

            if actual_sheet:
                if actual_sheet not in wb.sheetnames:
                    result["valid"] = False
                    result["errors"].append(f"Sheet '{actual_sheet}' 不存在")
                    wb.close()
                    return result
                sheet = wb[actual_sheet]
            else:
                sheet = wb.active

            # 使用统一解析器
            structure = self._parse_luban_sheet(sheet)

            # 将结构错误/警告合并到结果
            result["errors"].extend(structure.format_errors)
            result["warnings"].extend(structure.format_warnings)

            # 字段级检查
            if not structure.fields:
                result["warnings"].append("表没有定义字段")

            # 数据行值类型校验
            if structure.data_start_row > 0 and structure.fields:
                row_idx = 0
                for row in sheet.iter_rows(min_row=structure.data_start_row, values_only=True):
                    if all(c is None for c in row[1:len(structure.fields)+1]):
                        continue

                    for field in structure.fields:
                        col_idx = field["index"] + 1
                        if col_idx < len(row):
                            value = row[col_idx]
                            field_type = field["type"]

                            if value is not None:
                                error = self._validate_value(value, field_type, field["name"], row_idx)
                                if error:
                                    result["errors"].append(error)

                    row_idx += 1

            wb.close()

        except Exception as e:
            result["valid"] = False
            result["errors"].append(f"读取表失败: {e}")

        if result["errors"]:
            result["valid"] = False

        return result
    
    def _validate_value(self, value: Any, field_type: str, field_name: str, row_idx: int) -> Optional[str]:
        """验证单个值"""
        field_type_lower = field_type.lower() if field_type else ""
        
        # 跳过复杂类型验证
        if any(t in field_type_lower for t in ["list", "map", "vector", "array"]):
            return None
        
        # int 类型验证
        if field_type_lower == "int":
            if isinstance(value, str) and not value.lstrip('-').isdigit():
                return f"行 {row_idx}，字段 '{field_name}': 值 '{value}' 不是有效的整数"
        
        # float 类型验证
        if field_type_lower == "float":
            if isinstance(value, str):
                try:
                    float(value)
                except ValueError:
                    return f"行 {row_idx}，字段 '{field_name}': 值 '{value}' 不是有效的浮点数"
        
        # bool 类型验证
        if field_type_lower == "bool":
            if isinstance(value, str) and value.lower() not in ["true", "false", "1", "0"]:
                return f"行 {row_idx}，字段 '{field_name}': 值 '{value}' 不是有效的布尔值"
        
        return None
    
    def validate_all(self) -> Dict[str, Any]:
        """验证所有表"""
        tables = self.list_tables()
        results = {
            "total": len(tables),
            "valid": 0,
            "invalid": 0,
            "details": []
        }
        
        for table in tables:
            table_name = table["full_name"]
            result = self.validate_table(table_name)
            
            if result["valid"]:
                results["valid"] += 1
            else:
                results["invalid"] += 1
            
            results["details"].append({
                "table": table_name,
                "valid": result["valid"],
                "errors": result["errors"],
                "warnings": result["warnings"]
            })
        
        return results
    
    # ==================== Luban CLI 集成 ====================
    
    def gen(self, output_dir: str = None, luban_cmd: str = "dotnet run --project Luban.CLI") -> bool:
        """调用 Luban CLI 生成代码

        Args:
            output_dir: 输出目录
            luban_cmd: Luban CLI 命令
        """
        import subprocess
        import shutil

        # Pre-validation: 检查所有表的数据文件格式
        validate_result = self.validate_all()
        format_errors = []
        for detail in validate_result["details"]:
            # 只关注 format_errors（结构错误），不关注值类型错误
            for err in detail.get("errors", []):
                if any(kw in err for kw in ["缺少", "##type", "##var", "A 列", "类型定义"]):
                    format_errors.append(f"  {detail['table']}: {err}")

        if format_errors:
            print("✗ 数据文件格式检查失败，请先修复以下错误：")
            for err in format_errors:
                print(err)
            print("\n提示: 使用 'validate' 命令查看完整错误列表")
            return False
        
        # 检查 Luban.CLI 是否存在
        luban_cli_path = self.data_dir.parent / "Luban.CLI"
        if not luban_cli_path.exists():
            # 尝试其他常见位置
            possible_paths = [
                self.data_dir / "Luban.CLI",
                self.data_dir.parent.parent / "Luban.CLI",
            ]
            for p in possible_paths:
                if p.exists():
                    luban_cli_path = p
                    break
            else:
                print("错误: 未找到 Luban.CLI 目录")
                print("请确保 Luban.CLI 在项目目录中，或通过 --luban-cmd 指定完整命令")
                return False
        
        # 构建命令
        data_dir_str = str(self.data_dir)
        
        if output_dir:
            output_arg = f"-o {output_dir}"
        else:
            output_arg = ""
        
        cmd = f"{luban_cmd} -t all --conf {data_dir_str}/luban.conf {output_arg}"
        
        print(f"执行: {cmd}")
        print("-" * 50)
        
        try:
            result = subprocess.run(
                cmd,
                shell=True,
                cwd=str(luban_cli_path.parent) if luban_cli_path.exists() else ".",
                capture_output=True,
                text=True
            )
            
            if result.stdout:
                print(result.stdout)
            if result.stderr:
                print(result.stderr)
            
            if result.returncode == 0:
                print("-" * 50)
                print("✓ Luban 生成成功")
                return True
            else:
                print("-" * 50)
                print(f"✗ Luban 生成失败，返回码: {result.returncode}")
                return False
                
        except Exception as e:
            print(f"错误: 执行 Luban CLI 失败 - {e}")
            return False
    
    # ==================== 引用完整性检查 ====================
    
    def _build_type_index(self) -> Dict[str, List[str]]:
        """构建类型引用索引
        
        Returns:
            {被引用类型: [引用者列表]}
        """
        index = {}
        
        # 索引枚举
        enums = self.list_enums()
        for enum in enums:
            enum_name = enum["full_name"]
            index[enum_name] = []
        
        # 索引 Bean
        beans = self.list_beans()
        for bean in beans:
            bean_name = bean["full_name"]
            index[bean_name] = []
            
            # 检查字段类型引用
            for field in bean.get("fields", []):
                field_type = field.get("type", "")
                refs = self._extract_type_refs(field_type)
                for ref in refs:
                    if ref in index:
                        index[ref].append(f"{bean_name}.{field['name']}")

        # 索引表 — 优先通过 value_type 关联到已索引的 Bean，避免解析数据文件
        tables = self.list_tables()
        for table in tables:
            table_name = table["full_name"]
            value_type = table.get("value_type", "")
            if value_type and value_type in index:
                # 表的记录类型已通过 Bean 索引，无需重复解析数据文件
                continue
            # value_type 未在 Bean 中注册（如自动推导类型），尝试解析数据文件
            fields = self.list_fields(table_name)
            for field in fields:
                field_type = field.get("type", "")
                refs = self._extract_type_refs(field_type)
                for ref in refs:
                    if ref in index:
                        index[ref].append(f"{table_name}.{field['name']}")
        
        return index
    
    def _extract_type_refs(self, type_str: str) -> List[str]:
        """从类型字符串中提取引用的类型名"""
        refs = []

        # 移除容器类型
        type_str = type_str.strip()
        
        # 处理 list<xxx>, array<xxx>
        import re
        container_match = re.match(r'(list|array|vector)<(.+)>', type_str, re.IGNORECASE)
        if container_match:
            inner = container_match.group(2)
            refs.extend(self._extract_type_refs(inner))
        
        # 处理 map<k,v>
        map_match = re.match(r'map<(.+),\s*(.+)>', type_str, re.IGNORECASE)
        if map_match:
            refs.extend(self._extract_type_refs(map_match.group(1)))
            refs.extend(self._extract_type_refs(map_match.group(2)))
        
        # 检查是否是自定义类型（包含点或大写开头）
        if '.' in type_str or (type_str and type_str[0].isupper() and '<' not in type_str):
            # 不是基本类型
            basic_types = ['int', 'string', 'bool', 'float', 'long', 'double', 'byte', 'short']
            if type_str.lower() not in basic_types:
                refs.append(type_str)
        
        return refs
    
    def check_references(self, type_name: str) -> Dict[str, Any]:
        """检查类型的引用情况
        
        Args:
            type_name: 类型名称（枚举或Bean）
            
        Returns:
            {"type": "类型名", "referenced_by": ["引用者列表"], "can_delete": bool}
        """
        index = self._build_type_index()
        
        # 查找匹配的类型
        matched_type = None
        for t in index:
            if t == type_name or t.endswith("." + type_name):
                matched_type = t
                break
        
        if not matched_type:
            return {
                "type": type_name,
                "found": False,
                "referenced_by": [],
                "can_delete": False,
                "error": f"类型 {type_name} 不存在"
            }
        
        refs = index.get(matched_type, [])
        
        return {
            "type": matched_type,
            "found": True,
            "referenced_by": refs,
            "can_delete": len(refs) == 0
        }
    
    def delete_bean_safe(self, bean_name: str, force: bool = False) -> bool:
        """安全删除 Bean（检查引用）
        
        Args:
            bean_name: Bean 名称
            force: 强制删除
        """
        ref_check = self.check_references(bean_name)
        
        if not ref_check["found"]:
            print(f"错误: {ref_check['error']}")
            return False
        
        if ref_check["referenced_by"]:
            print(f"警告: Bean {bean_name} 被以下位置引用:")
            for ref in ref_check["referenced_by"]:
                print(f"  - {ref}")
            
            if not force:
                print("\n使用 --force 强制删除")
                return False
        
        return self.delete_bean(bean_name)
    
    def delete_enum_safe(self, enum_name: str, force: bool = False) -> bool:
        """安全删除枚举（检查引用）
        
        Args:
            enum_name: 枚举名称
            force: 强制删除
        """
        ref_check = self.check_references(enum_name)
        
        if not ref_check["found"]:
            print(f"错误: {ref_check['error']}")
            return False
        
        if ref_check["referenced_by"]:
            print(f"警告: 枚举 {enum_name} 被以下位置引用:")
            for ref in ref_check["referenced_by"]:
                print(f"  - {ref}")
            
            if not force:
                print("\n使用 --force 强制删除")
                return False
        
        return self.delete_enum(enum_name)
    
    # ==================== 配置模板 ====================
    
    TEMPLATES = {
        "item": {
            "name": "道具表",
            "fields": [
                {"name": "id", "type": "int", "comment": "道具ID", "group": "c"},
                {"name": "name", "type": "string", "comment": "道具名称", "group": "c"},
                {"name": "type", "type": "int", "comment": "道具类型", "group": "c"},
                {"name": "quality", "type": "int", "comment": "品质(1-5)", "group": "c"},
                {"name": "max_stack", "type": "int", "comment": "最大堆叠", "group": "c"},
                {"name": "price", "type": "int", "comment": "出售价格", "group": "s"},
                {"name": "description", "type": "string", "comment": "描述", "group": "c"},
            ]
        },
        "skill": {
            "name": "技能表",
            "fields": [
                {"name": "id", "type": "int", "comment": "技能ID", "group": "c"},
                {"name": "name", "type": "string", "comment": "技能名称", "group": "c"},
                {"name": "type", "type": "int", "comment": "技能类型", "group": "c"},
                {"name": "target_type", "type": "int", "comment": "目标类型", "group": "s"},
                {"name": "cooldown", "type": "int", "comment": "冷却时间(秒)", "group": "s"},
                {"name": "mp_cost", "type": "int", "comment": "法力消耗", "group": "s"},
                {"name": "description", "type": "string", "comment": "描述", "group": "c"},
            ]
        },
        "hero": {
            "name": "英雄表",
            "fields": [
                {"name": "id", "type": "int", "comment": "英雄ID", "group": "c"},
                {"name": "name", "type": "string", "comment": "英雄名称", "group": "c"},
                {"name": "quality", "type": "int", "comment": "品质", "group": "c"},
                {"name": "hp", "type": "int", "comment": "生命值", "group": "s"},
                {"name": "attack", "type": "int", "comment": "攻击力", "group": "s"},
                {"name": "defense", "type": "int", "comment": "防御力", "group": "s"},
                {"name": "speed", "type": "int", "comment": "速度", "group": "s"},
                {"name": "skill_ids", "type": "list<int>", "comment": "技能列表", "group": "s"},
                {"name": "description", "type": "string", "comment": "描述", "group": "c"},
            ]
        },
        "reward": {
            "name": "奖励表",
            "fields": [
                {"name": "id", "type": "int", "comment": "奖励ID", "group": "c"},
                {"name": "name", "type": "string", "comment": "奖励名称", "group": "c"},
                {"name": "items", "type": "list<test.RewardItem>", "comment": "奖励列表", "group": "s"},
            ]
        },
        "level": {
            "name": "等级表",
            "fields": [
                {"name": "level", "type": "int", "comment": "等级", "group": "c"},
                {"name": "exp", "type": "int", "comment": "所需经验", "group": "s"},
                {"name": "reward_id", "type": "int", "comment": "升级奖励", "group": "s"},
            ]
        },
    }
    
    def list_templates(self) -> List[Dict[str, Any]]:
        """列出所有可用模板"""
        return [
            {"name": name, "description": data["name"], "field_count": len(data["fields"])}
            for name, data in self.TEMPLATES.items()
        ]
    
    def create_from_template(self, template_name: str, table_name: str, 
                             input_file: str = None, module: str = "test") -> bool:
        """从模板创建表
        
        Args:
            template_name: 模板名称
            table_name: 表名称（不含模块）
            input_file: 输入文件名
            module: 模块名
        """
        if template_name not in self.TEMPLATES:
            print(f"错误: 模板 '{template_name}' 不存在")
            print(f"可用模板: {', '.join(self.TEMPLATES.keys())}")
            return False
        
        template = self.TEMPLATES[template_name]
        full_name = f"{module}.{table_name}"
        
        if not input_file:
            input_file = f"{table_name.lower()}.xlsx"
        
        # 调用 add_table
        print(f"从模板 '{template_name}' 创建表 {full_name}")
        return self.add_table(
            full_name=full_name,
            fields=template["fields"],
            input_file=input_file,
            comment=template["name"],
            index=template["fields"][0]["name"],  # 第一个字段作为主键
            auto_import=False
        )
    
    # ==================== 数据迁移 ====================
    
    def rename_table(self, old_name: str, new_name: str, 
                     migrate_data: bool = False) -> bool:
        """重命名表
        
        Args:
            old_name: 原表名
            new_name: 新表名
            migrate_data: 是否迁移数据文件
        """
        if not self.tables_file.exists():
            print(f"错误: 文件不存在 {self.tables_file}")
            return False
        
        wb = openpyxl.load_workbook(self.tables_file)
        sheet = wb.active
        
        # 找到表定义行
        target_row = None
        for i, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
            full_name = row[1]
            if full_name == old_name or (full_name and full_name.endswith("." + old_name)):
                target_row = i
                break
        
        if not target_row:
            print(f"错误: 未找到表 {old_name}")
            wb.close()
            return False
        
        # 更新表名
        sheet.cell(row=target_row, column=2, value=new_name)
        
        wb.save(self.tables_file)
        wb.close()
        
        print(f"✓ 已重命名表: {old_name} -> {new_name}")
        
        # 迁移数据文件
        if migrate_data:
            # 获取原输入文件
            table = self.get_table(new_name)
            if table:
                old_input = table.get("input", "")
                if old_input:
                    old_path = self.data_dir / str(old_input)
                    if old_path.exists():
                        # 构建新文件名
                        new_input = old_input.replace(old_name.split(".")[-1], new_name.split(".")[-1])
                        new_path = self.data_dir / new_input
                        
                        import shutil
                        shutil.move(str(old_path), str(new_path))
                        
                        # 更新 input 字段
                        wb = openpyxl.load_workbook(self.tables_file)
                        sheet = wb.active
                        sheet.cell(row=target_row, column=5, value=new_input)  # E列 = input
                        wb.save(self.tables_file)
                        wb.close()
                        
                        print(f"✓ 已迁移数据文件: {old_input} -> {new_input}")
        
        return True
    
    def copy_table(self, source_name: str, target_name: str, 
                   copy_data: bool = False) -> bool:
        """复制表
        
        Args:
            source_name: 源表名
            target_name: 目标表名
            copy_data: 是否复制数据文件
        """
        # 获取源表信息
        source_table = self.get_table(source_name)
        if not source_table:
            print(f"错误: 源表 {source_name} 不存在")
            return False
        
        source_input = source_table.get("input", "")
        
        # 复制数据文件
        if copy_data and source_input:
            source_path = self.data_dir / str(source_input)
            if source_path.exists():
                target_input = source_input.replace(
                    source_name.split(".")[-1], 
                    target_name.split(".")[-1]
                )
                target_path = self.data_dir / target_input
                
                import shutil
                shutil.copy2(str(source_path), str(target_path))
                
                print(f"✓ 已复制数据文件: {source_input} -> {target_input}")
        
        # 在 __tables__.xlsx 中添加新表定义
        wb = openpyxl.load_workbook(self.tables_file)
        sheet = wb.active
        
        # 找到源表行
        source_row = None
        for i, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
            if row[1] == source_name or (row[1] and row[1].endswith("." + source_name)):
                source_row = i
                break
        
        if source_row:
            # 复制行
            new_row = sheet.max_row + 1
            for col in range(1, sheet.max_column + 1):
                value = sheet.cell(row=source_row, column=col).value
                if col == 2:  # 表名列
                    value = target_name
                if col == 5 and copy_data:  # E列 = input 列
                    value = target_input if source_input else value
                sheet.cell(row=new_row, column=col, value=value)
            
            wb.save(self.tables_file)
            print(f"✓ 已复制表定义: {source_name} -> {target_name}")
        
        wb.close()
        return True
    
    # ==================== 差异对比 ====================
    
    def diff_tables(self, table1_name: str, table2_name: str) -> Dict[str, Any]:
        """对比两个表的差异
        
        Args:
            table1_name: 表1名称
            table2_name: 表2名称
        """
        fields1 = self.list_fields(table1_name)
        fields2 = self.list_fields(table2_name)
        
        fields1_map = {f["name"]: f for f in fields1}
        fields2_map = {f["name"]: f for f in fields2}
        
        result = {
            "table1": table1_name,
            "table2": table2_name,
            "added_fields": [],      # 表2新增的字段
            "removed_fields": [],    # 表2删除的字段
            "modified_fields": [],   # 类型或注释变更的字段
            "identical_fields": []   # 相同的字段
        }
        
        # 检查表1的字段
        for name, f1 in fields1_map.items():
            if name not in fields2_map:
                result["removed_fields"].append(name)
            else:
                f2 = fields2_map[name]
                changes = []
                if f1.get("type") != f2.get("type"):
                    changes.append(f"type: {f1.get('type')} -> {f2.get('type')}")
                if f1.get("comment") != f2.get("comment"):
                    changes.append(f"comment changed")
                if f1.get("group") != f2.get("group"):
                    changes.append(f"group: {f1.get('group')} -> {f2.get('group')}")
                
                if changes:
                    result["modified_fields"].append({
                        "name": name,
                        "changes": changes
                    })
                else:
                    result["identical_fields"].append(name)
        
        # 检查表2新增的字段
        for name in fields2_map:
            if name not in fields1_map:
                result["added_fields"].append(name)
        
        return result
    
    def diff_with_json(self, table_name: str, json_file: str) -> Dict[str, Any]:
        """对比表与 JSON 文件的差异
        
        Args:
            table_name: 表名称
            json_file: JSON 文件路径
        """
        import json as json_module
        
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                json_data = json_module.load(f)
        except Exception as e:
            return {"error": f"读取 JSON 文件失败: {e}"}
        
        json_fields = json_data.get("fields", [])
        current_fields = self.list_fields(table_name)
        
        json_map = {f["name"]: f for f in json_fields}
        current_map = {f["name"]: f for f in current_fields}
        
        result = {
            "table": table_name,
            "json_file": json_file,
            "added_fields": [],
            "removed_fields": [],
            "modified_fields": [],
            "identical_fields": []
        }
        
        for name, cf in current_map.items():
            if name not in json_map:
                result["removed_fields"].append(name)
            else:
                jf = json_map[name]
                changes = []
                if cf.get("type") != jf.get("type"):
                    changes.append(f"type: {cf.get('type')} -> {jf.get('type')}")
                if cf.get("comment") != jf.get("comment"):
                    changes.append("comment changed")
                
                if changes:
                    result["modified_fields"].append({"name": name, "changes": changes})
                else:
                    result["identical_fields"].append(name)
        
        for name in json_map:
            if name not in current_map:
                result["added_fields"].append(name)
        
        return result
    
    # ==================== 自动导入表 ====================
    
    def list_auto_import_tables(self) -> List[Dict[str, Any]]:
        """列出所有自动导入的表（文件名以 # 开头）"""
        auto_tables = []
        
        # 扫描数据目录下所有以 # 开头的 Excel 文件
        for file_path in self.data_dir.rglob('#*'):
            if file_path.suffix.lower() in ['.xlsx', '.xls', '.csv']:
                # 解析表名
                file_stem = file_path.stem  # 文件名（不含扩展名）
                
                # 解析格式: #TableName 或 #TableName-表注释
                if '-' in file_stem:
                    parts = file_stem[1:].split('-', 1)
                    value_type = parts[0]
                    comment = parts[1] if len(parts) > 1 else ""
                else:
                    value_type = file_stem[1:]
                    comment = ""
                
                # 计算命名空间（子目录）
                rel_path = file_path.relative_to(self.data_dir)
                if len(rel_path.parts) > 1:
                    # 有子目录，作为命名空间
                    namespace = '.'.join(rel_path.parts[:-1])
                    full_name = f"{namespace}.Tb{value_type}"
                    full_value_type = f"{namespace}.{value_type}"
                else:
                    full_name = f"Tb{value_type}"
                    full_value_type = value_type
                
                auto_tables.append({
                    "file": file_path.relative_to(self.data_dir).as_posix(),
                    "full_name": full_name,
                    "value_type": full_value_type,
                    "comment": comment,
                    "mode": "map"
                })
        
        return auto_tables
    
    def create_auto_import_table(self, file_name: str, fields: str = None) -> bool:
        """创建自动导入表（创建 # 前缀的 Excel 文件）
        
        Args:
            file_name: 文件名（如 #Item 或 #Item-道具表）
            fields: 字段定义，格式 name:type:comment
        """
        # 解析文件名
        if not file_name.startswith('#'):
            file_name = '#' + file_name
        
        if '-' in file_name:
            parts = file_name.split('-', 1)
            value_type = parts[0][1:]  # 去掉 #
            comment = parts[1]
        else:
            value_type = file_name[1:]
            comment = ""
        
        # 创建 Excel 文件
        excel_path = self.data_dir / f"{file_name}.xlsx"
        
        if excel_path.exists():
            print(f"错误: 文件已存在 {excel_path}")
            return False
        
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Sheet1"
        
        # 写入头部
        sheet.cell(row=1, column=1, value="##var")
        sheet.cell(row=2, column=1, value="##type")
        sheet.cell(row=3, column=1, value="##")
        
        # 解析字段并写入
        if fields:
            field_list = []
            for f in fields.split(','):
                parts = f.split(':')
                field_list.append({
                    "name": parts[0] if len(parts) > 0 else "",
                    "type": parts[1] if len(parts) > 1 else "int",
                    "comment": parts[2] if len(parts) > 2 else ""
                })
            
            for i, field in enumerate(field_list):
                col = i + 2
                sheet.cell(row=1, column=col, value=field["name"])
                sheet.cell(row=2, column=col, value=field["type"])
                sheet.cell(row=3, column=col, value=field["comment"])
        
        # 写入注释行
        if comment:
            sheet.cell(row=3, column=1, value=f"## {comment}")
        
        wb.save(excel_path)
        wb.close()
        
        print(f"✓ 已创建自动导入表: {excel_path}")
        print(f"  表名: Tb{value_type}")
        print(f"  记录类型: {value_type}")
        return True
    
    # ==================== 常量别名 ====================
    
    def list_constalias(self) -> List[Dict[str, Any]]:
        """列出所有常量别名"""
        # 常量别名存储在 __enums__.xlsx 或单独的 __constalias__.xlsx
        constalias_file = self.data_dir / "__constalias__.xlsx"
        
        if not constalias_file.exists():
            return []
        
        wb = openpyxl.load_workbook(constalias_file)
        sheet = wb.active
        
        aliases = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:  # name
                aliases.append({
                    "name": row[0],
                    "value": self._cell(row, 1),
                    "comment": self._cell(row, 2)
                })
        
        wb.close()
        return aliases
    
    def add_constalias(self, name: str, value: str, comment: str = "") -> bool:
        """添加常量别名"""
        constalias_file = self.data_dir / "__constalias__.xlsx"
        
        # 如果文件不存在，创建
        if not constalias_file.exists():
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "constalias"
            sheet.cell(row=1, column=1, value="name")
            sheet.cell(row=1, column=2, value="value")
            sheet.cell(row=1, column=3, value="comment")
        else:
            wb = openpyxl.load_workbook(constalias_file)
            sheet = wb.active
        
        # 检查是否已存在
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == name:
                print(f"错误: 常量别名 '{name}' 已存在")
                wb.close()
                return False
        
        # 添加新行
        new_row = sheet.max_row + 1
        sheet.cell(row=new_row, column=1, value=name)
        sheet.cell(row=new_row, column=2, value=value)
        sheet.cell(row=new_row, column=3, value=comment)
        
        wb.save(constalias_file)
        wb.close()
        
        print(f"✓ 已添加常量别名: {name} = {value}")
        return True
    
    def delete_constalias(self, name: str) -> bool:
        """删除常量别名"""
        constalias_file = self.data_dir / "__constalias__.xlsx"
        
        if not constalias_file.exists():
            print(f"错误: 常量别名文件不存在")
            return False
        
        wb = openpyxl.load_workbook(constalias_file)
        sheet = wb.active
        
        # 查找并删除
        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] == name:
                sheet.delete_rows(i)
                wb.save(constalias_file)
                wb.close()
                print(f"✓ 已删除常量别名: {name}")
                return True
        
        wb.close()
        print(f"错误: 未找到常量别名 '{name}'")
        return False
    
    def resolve_constalias(self, name: str) -> Optional[str]:
        """解析常量别名的值"""
        aliases = self.list_constalias()
        for alias in aliases:
            if alias["name"] == name:
                return alias["value"]
        return None
    
    # ==================== 数据标签过滤 ====================
    
    def list_data_tags(self, table_name: str, sheet_name: str = None) -> List[Dict[str, Any]]:
        """列出表中所有使用的数据标签"""
        result = self._get_table_excel_path(table_name, sheet_name)
        if not result:
            return []
        
        excel_path, actual_sheet = result
        wb = openpyxl.load_workbook(excel_path)
        
        if actual_sheet:
            sheet = wb[actual_sheet]
        else:
            sheet = wb.active
        
        # 找到数据起始行
        data_start = self._get_data_start_row(sheet)
        
        tags = set()
        tagged_rows = []
        
        for i, row in enumerate(sheet.iter_rows(min_row=data_start, values_only=True), start=0):
            # 第一列是标签列（如果有的话）
            first_cell = row[0] if row else None
            if first_cell and isinstance(first_cell, str) and first_cell.startswith('#'):
                # ## 表示永久注释，其他 #tag 表示标签
                if first_cell != "##":
                    tag = first_cell[1:]  # 去掉 #
                    tags.add(tag)
                    tagged_rows.append({"row": i, "tag": tag})
        
        wb.close()
        
        return {
            "table": table_name,
            "tags": list(tags),
            "tagged_rows": tagged_rows
        }
    
    def tag_row(self, table_name: str, row_index: int, tag: str, 
                sheet_name: str = None) -> bool:
        """给数据行添加标签
        
        Args:
            table_name: 表名称
            row_index: 行索引
            tag: 标签名（不含 #）
        """
        result = self._get_table_excel_path(table_name, sheet_name)
        if not result:
            return False
        
        excel_path, actual_sheet = result
        wb = openpyxl.load_workbook(excel_path)
        
        if actual_sheet:
            sheet = wb[actual_sheet]
        else:
            sheet = wb.active
        
        data_start = self._get_data_start_row(sheet)
        target_row = data_start + row_index
        
        if target_row > sheet.max_row:
            print(f"错误: 行索引 {row_index} 超出范围")
            wb.close()
            return False
        
        # 在第一列添加标签
        current_value = sheet.cell(row=target_row, column=1).value
        if current_value is None or current_value == "":
            sheet.cell(row=target_row, column=1, value=f"#{tag}")
        elif isinstance(current_value, str) and current_value.startswith("#"):
            # 已有标签，追加
            sheet.cell(row=target_row, column=1, value=f"{current_value},{tag}")
        else:
            sheet.cell(row=target_row, column=1, value=f"#{tag}")
        
        wb.save(excel_path)
        wb.close()
        
        print(f"✓ 已为行 {row_index} 添加标签: {tag}")
        return True
    
    def untag_row(self, table_name: str, row_index: int, 
                  sheet_name: str = None) -> bool:
        """移除数据行的标签"""
        result = self._get_table_excel_path(table_name, sheet_name)
        if not result:
            return False
        
        excel_path, actual_sheet = result
        wb = openpyxl.load_workbook(excel_path)
        
        if actual_sheet:
            sheet = wb[actual_sheet]
        else:
            sheet = wb.active
        
        data_start = self._get_data_start_row(sheet)
        target_row = data_start + row_index
        
        if target_row > sheet.max_row:
            print(f"错误: 行索引 {row_index} 超出范围")
            wb.close()
            return False
        
        # 清除标签
        sheet.cell(row=target_row, column=1, value=None)
        
        wb.save(excel_path)
        wb.close()
        
        print(f"✓ 已移除行 {row_index} 的标签")
        return True
    
    # ==================== 字段变体 ====================
    
    def list_field_variants(self, table_name: str, field_name: str,
                            sheet_name: str = None) -> List[str]:
        """列出字段的变体"""
        result = self._get_table_excel_path(table_name, sheet_name)
        if not result:
            return []
        
        excel_path, actual_sheet = result
        wb = openpyxl.load_workbook(excel_path)
        
        if actual_sheet:
            sheet = wb[actual_sheet]
        else:
            sheet = wb.active
        
        # 在标题行查找变体字段
        var_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        
        variants = []
        base_col = None
        
        for i, cell in enumerate(var_row[1:], start=2):
            if cell == field_name:
                base_col = i
            elif isinstance(cell, str) and cell.startswith(f"{field_name}@"):
                variant = cell.split('@')[1]
                variants.append(variant)
        
        wb.close()
        
        return {
            "field": field_name,
            "base_column": base_col,
            "variants": variants
        }
    
    def add_field_variant(self, table_name: str, field_name: str, variant: str,
                          sheet_name: str = None) -> bool:
        """添加字段变体
        
        Args:
            table_name: 表名称
            field_name: 字段名
            variant: 变体名（如 zh, en）
        """
        result = self._get_table_excel_path(table_name, sheet_name)
        if not result:
            return False
        
        excel_path, actual_sheet = result
        wb = openpyxl.load_workbook(excel_path)
        
        if actual_sheet:
            sheet = wb[actual_sheet]
        else:
            sheet = wb.active
        
        # 找到原字段列
        var_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        type_row = list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        
        base_col = None
        base_type = None
        
        for i, cell in enumerate(var_row[1:], start=2):
            if cell == field_name:
                base_col = i
                base_type = type_row[i-1] if i-1 < len(type_row) else None
                break
        
        if not base_col:
            print(f"错误: 未找到字段 '{field_name}'")
            wb.close()
            return False
        
        # 检查变体是否已存在
        variant_name = f"{field_name}@{variant}"
        for cell in var_row[base_col:]:
            if cell == variant_name:
                print(f"错误: 变体 '{variant}' 已存在")
                wb.close()
                return False
        
        # 在字段后面插入新列
        sheet.insert_cols(base_col + 1)
        
        # 设置变体标题和类型
        sheet.cell(row=1, column=base_col + 1, value=variant_name)
        sheet.cell(row=2, column=base_col + 1, value=base_type)
        
        wb.save(excel_path)
        wb.close()
        
        print(f"✓ 已添加字段变体: {field_name}@{variant}")
        return True
    
    # ==================== 多行结构列表 ====================
    
    def set_multi_row_field(self, table_name: str, field_name: str,
                            enable: bool = True, sheet_name: str = None) -> bool:
        """设置字段为多行结构列表
        
        Args:
            table_name: 表名称
            field_name: 字段名
            enable: True 添加 * 前缀，False 移除
        """
        result = self._get_table_excel_path(table_name, sheet_name)
        if not result:
            return False
        
        excel_path, actual_sheet = result
        wb = openpyxl.load_workbook(excel_path)
        
        if actual_sheet:
            sheet = wb[actual_sheet]
        else:
            sheet = wb.active
        
        # 找到字段列
        var_row_num = 1
        for col in range(2, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col).value
            if cell_value == field_name or cell_value == f"*{field_name}":
                if enable:
                    new_name = f"*{field_name}" if not field_name.startswith("*") else field_name
                else:
                    new_name = field_name.lstrip("*")
                
                sheet.cell(row=1, column=col, value=new_name)
                wb.save(excel_path)
                wb.close()
                
                status = "启用" if enable else "禁用"
                print(f"✓ 已{status}多行结构: {new_name}")
                return True
        
        wb.close()
        print(f"错误: 未找到字段 '{field_name}'")
        return False
    
    # ==================== 可空类型支持 ====================
    
    def is_nullable_type(self, type_str: str) -> bool:
        """检查类型是否为可空类型"""
        return type_str.endswith('?')
    
    def get_base_type(self, type_str: str) -> str:
        """获取可空类型的基础类型"""
        if type_str.endswith('?'):
            return type_str[:-1]
        return type_str
    
    def validate_nullable_value(self, value: Any, type_str: str) -> bool:
        """验证可空类型的值"""
        if value is None or value == "":
            return True  # 可空类型允许空值
        
        base_type = self.get_base_type(type_str)
        return self._validate_value(value, base_type, "", 0) is None
    
    # ==================== 容器类型支持 ====================
    
    CONTAINER_TYPES = ['array', 'list', 'set', 'map']
    
    def parse_container_type(self, type_str: str) -> Dict[str, Any]:
        """解析容器类型
        
        Returns:
            {"container": "list", "element_type": "int"} 或
            {"container": "map", "key_type": "int", "value_type": "string"}
        """
        import re
        
        # 匹配 array<int>, list<int>, set<int>
        match = re.match(r'(array|list|set)<(.+)>', type_str, re.IGNORECASE)
        if match:
            return {
                "container": match.group(1).lower(),
                "element_type": match.group(2)
            }
        
        # 匹配 map<int,string>
        match = re.match(r'map<(.+),\s*(.+)>', type_str, re.IGNORECASE)
        if match:
            return {
                "container": "map",
                "key_type": match.group(1),
                "value_type": match.group(2)
            }
        
        return {"container": None, "type": type_str}
    
    def get_type_info(self, type_str: str) -> Dict[str, Any]:
        """获取类型的完整信息"""
        info = {
            "original": type_str,
            "is_nullable": self.is_nullable_type(type_str),
            "is_container": False,
            "is_enum": False,
            "is_bean": False,
            "base_type": None,
            "container_info": None
        }
        
        # 处理可空
        base = self.get_base_type(type_str)
        info["base_type"] = base
        
        # 检查容器类型
        container_info = self.parse_container_type(base)
        if container_info["container"]:
            info["is_container"] = True
            info["container_info"] = container_info
            return info
        
        # 检查基本类型
        basic_types = ['bool', 'byte', 'short', 'int', 'long', 'float', 'double', 'string', 'text', 'datetime']
        if base.lower() in basic_types:
            return info
        
        # 检查枚举
        enums = self.list_enums()
        for enum in enums:
            if enum["full_name"] == base or enum["full_name"].endswith("." + base):
                info["is_enum"] = True
                return info
        
        # 检查 Bean
        beans = self.list_beans()
        for bean in beans:
            if bean["full_name"] == base or bean["full_name"].endswith("." + base):
                info["is_bean"] = True
                return info
        
        return info

    def list_all_types(self, category: str = "all") -> Dict[str, Any]:
        """列出所有可用类型

        Args:
            category: 类型类别筛选 - basic(基本类型), container(容器类型), enum(枚举), bean(Bean), all(全部)

        Returns:
            包含各类型的详细信息
        """
        result = {
            "category": category,
            "types": []
        }

        # 基本类型
        basic_types = [
            {"name": "bool", "description": "布尔值", "example": "true/false"},
            {"name": "byte", "description": "8位有符号整数", "example": "-128 ~ 127"},
            {"name": "short", "description": "16位有符号整数", "example": "-32768 ~ 32767"},
            {"name": "int", "description": "32位有符号整数", "example": "常用数字类型"},
            {"name": "long", "description": "64位有符号整数", "example": "大整数"},
            {"name": "float", "description": "32位浮点数", "example": "3.14"},
            {"name": "double", "description": "64位浮点数", "example": "高精度小数"},
            {"name": "string", "description": "字符串", "example": "文本内容"},
            {"name": "text", "description": "长文本", "example": "大段文字"},
            {"name": "datetime", "description": "日期时间", "example": "2024-01-01 12:00:00"}
        ]

        # 容器类型模板
        container_templates = [
            {"name": "array<T>", "description": "定长数组", "example": "array<int>"},
            {"name": "list<T>", "description": "变长列表", "example": "list<string>"},
            {"name": "set<T>", "description": "集合（去重）", "example": "set<int>"},
            {"name": "map<K,V>", "description": "键值对映射", "example": "map<int,string>"}
        ]

        if category in ["basic", "all"]:
            for t in basic_types:
                result["types"].append({
                    "category": "basic",
                    "name": t["name"],
                    "description": t["description"],
                    "example": t["example"],
                    "nullable_version": f"{t['name']}?"
                })

        if category in ["container", "all"]:
            for t in container_templates:
                result["types"].append({
                    "category": "container",
                    "name": t["name"],
                    "description": t["description"],
                    "example": t["example"],
                    "note": "T/K/V 可以是任何基本类型、枚举或Bean"
                })

        if category in ["enum", "all"]:
            enums = self.list_enums()
            for enum in enums:
                result["types"].append({
                    "category": "enum",
                    "name": enum["full_name"],
                    "description": enum.get("comment", ""),
                    "items_count": len(enum.get("items", [])),
                    "items": [{"name": item["name"], "value": item["value"], "alias": item.get("alias", "")}
                             for item in enum.get("items", [])[:5]]  # 最多显示5个
                })

        if category in ["bean", "all"]:
            beans = self.list_beans()
            for bean in beans:
                result["types"].append({
                    "category": "bean",
                    "name": bean["full_name"],
                    "description": bean.get("comment", ""),
                    "parent": bean.get("parent", ""),
                    "fields_count": len(bean.get("fields", []))
                })

        return result

    def validate_type(self, type_str: str) -> Dict[str, Any]:
        """验证类型字符串是否有效

        Args:
            type_str: 类型字符串，如 "list<int>", "test.EQuality"

        Returns:
            验证结果，包含 is_valid 和详细信息
        """
        if not type_str or not type_str.strip():
            return {"is_valid": False, "error": "类型字符串为空"}

        type_str = type_str.strip()
        info = self.get_type_info(type_str)

        result = {
            "type": type_str,
            "is_valid": True,
            "info": info
        }

        # 检查是否为已知类型
        is_known = (
            info.get("is_container") or
            (info.get("base_type") and info.get("base_type").lower() in [
                'bool', 'byte', 'short', 'int', 'long', 'float', 'double', 'string', 'text', 'datetime'
            ]) or
            info.get("is_enum") or
            info.get("is_bean")
        )

        if not is_known:
            result["is_valid"] = False
            result["error"] = f"未知类型: {type_str}"
            result["suggestions"] = self._get_type_suggestions(type_str)

        return result

    def _get_type_suggestions(self, type_str: str) -> List[str]:
        """获取类型建议（当类型无效时）"""
        suggestions = []

        # 检查是否是缺少模块名的枚举/Bean
        enums = self.list_enums()
        beans = self.list_beans()

        # 查找相似名称
        for enum in enums:
            if type_str.lower() in enum["full_name"].lower() or enum["full_name"].lower().endswith(type_str.lower()):
                suggestions.append(enum["full_name"])

        for bean in beans:
            if type_str.lower() in bean["full_name"].lower() or bean["full_name"].lower().endswith(type_str.lower()):
                suggestions.append(bean["full_name"])

        return suggestions[:5]  # 最多返回5个建议

    def suggest_type_for_field(self, field_name: str, context: str = "general") -> Dict[str, Any]:
        """根据字段名建议类型

        Args:
            field_name: 字段名，如 'item_id', 'name', 'quality'
            context: 上下文类型 - item(道具), skill(技能), monster(怪物), quest(任务), general(通用)

        Returns:
            建议的类型列表，按匹配度排序
        """
        field_name_lower = field_name.lower()
        suggestions = []

        # 通用模式匹配
        patterns = {
            # ID 相关
            r".*_?id$": ["int", "long", "string"],
            r"^id$": ["int", "long"],

            # 名称相关
            r".*name.*": ["string"],
            r".*desc.*": ["string", "text"],
            r".*title.*": ["string"],

            # 数量/数值
            r".*count$": ["int", "short"],
            r".*num$": ["int", "short"],
            r".*amount$": ["int", "long"],
            r".*max$": ["int", "short"],
            r".*min$": ["int", "short"],

            # 品质/等级
            r".*quality$": ["int"],
            r".*level$": ["int", "short"],
            r".*grade$": ["int", "short"],
            r".*rank$": ["int", "short"],
            r".*tier$": ["int", "short"],

            # 价格/货币
            r".*price$": ["int", "long"],
            r".*cost$": ["int", "long"],
            r".*gold$": ["int", "long"],
            r".*diamond$": ["int", "long"],

            # 概率/比率
            r".*rate$": ["float", "double"],
            r".*ratio$": ["float", "double"],
            r".*prob$": ["float", "double"],
            r".*chance$": ["float", "double"],

            # 时间
            r".*time$": ["int", "long", "datetime"],
            r".*duration$": ["int", "long"],
            r".*cd$": ["int", "float"],
            r".*cooldown$": ["int", "float"],

            # 布尔标志
            r"is_.*": ["bool"],
            r"has_.*": ["bool"],
            r"can_.*": ["bool"],
            r"enable.*": ["bool"],
            r".*_flag$": ["bool"],

            # 资源路径
            r".*icon$": ["string"],
            r".*image$": ["string"],
            r".*model$": ["string"],
            r".*prefab$": ["string"],
            r".*asset$": ["string"],
            r".*path$": ["string"],
            r".*url$": ["string"],

            # 列表/数组（复数形式）
            r".*s$": ["list<int>", "array<int>"],
            r".*_list$": ["list<int>", "list<string>"],
            r".*_ids$": ["list<int>", "list<long>"],
            r".*drops$": ["list<int>", "list<DropItem>"],
            r".*rewards$": ["list<RewardItem>", "list<int>"],

            # 坐标/位置
            r".*pos.*": ["float", "double"],
            r".*coord.*": ["float", "double"],
            r".*x$": ["float", "int"],
            r".*y$": ["float", "int"],
            r".*z$": ["float", "int"],
        }

        import re
        for pattern, types in patterns.items():
            if re.match(pattern, field_name_lower):
                for t in types:
                    suggestions.append({
                        "type": t,
                        "confidence": "high",
                        "reason": f"字段名 '{field_name}' 匹配模式 '{pattern}'"
                    })

        # 根据上下文调整建议
        context_adjustments = {
            "item": {
                "type": ["int", "EItemType"],
                "quality": ["int", "EQuality"],
                "bind": ["bool"],
                "stackable": ["bool"],
            },
            "skill": {
                "type": ["int", "ESkillType"],
                "target": ["int", "ETargetType"],
                "damage": ["int", "float"],
                "mp_cost": ["int"],
                "cd": ["float", "int"],
            },
            "monster": {
                "type": ["int", "EMonsterType"],
                "race": ["int", "ERace"],
                "hp": ["int", "long"],
                "atk": ["int"],
                "def": ["int"],
                "drops": ["list<int>", "list<DropItem>"],
            },
            "quest": {
                "type": ["int", "EQuestType"],
                "status": ["int", "EQuestStatus"],
                "npc_id": ["int"],
                "reward": ["list<int>", "list<RewardItem>"],
            }
        }

        if context in context_adjustments:
            for key, types in context_adjustments[context].items():
                if key in field_name_lower:
                    for t in types:
                        suggestions.insert(0, {
                            "type": t,
                            "confidence": "very_high",
                            "reason": f"{context} 上下文中的 '{key}' 字段"
                        })

        # 如果没有匹配，返回通用建议
        if not suggestions:
            suggestions = [
                {"type": "int", "confidence": "medium", "reason": "通用数字类型"},
                {"type": "string", "confidence": "medium", "reason": "通用文本类型"},
                {"type": "bool", "confidence": "low", "reason": "通用布尔类型"}
            ]

        # 去重
        seen = set()
        unique_suggestions = []
        for s in suggestions:
            if s["type"] not in seen:
                seen.add(s["type"])
                unique_suggestions.append(s)

        return {
            "field_name": field_name,
            "context": context,
            "suggestions": unique_suggestions[:5]  # 最多返回5个建议
        }

    def search_types(self, keyword: str, category: str = "all") -> Dict[str, Any]:
        """搜索类型

        Args:
            keyword: 搜索关键词
            category: 搜索类别 - enum(枚举), bean(Bean), all(全部)

        Returns:
            匹配的类型列表
        """
        keyword_lower = keyword.lower()
        results = []

        if category in ["enum", "all"]:
            enums = self.list_enums()
            for enum in enums:
                match_score = 0
                enum_comment = enum.get("comment") or ""
                if keyword_lower in enum["full_name"].lower():
                    match_score += 3
                if keyword_lower in enum_comment.lower():
                    match_score += 2
                for item in enum.get("items", []):
                    if keyword_lower in item.get("name", "").lower():
                        match_score += 1
                        break

                if match_score > 0:
                    results.append({
                        "category": "enum",
                        "name": enum["full_name"],
                        "comment": enum_comment,
                        "match_score": match_score
                    })

        if category in ["bean", "all"]:
            beans = self.list_beans()
            for bean in beans:
                match_score = 0
                bean_comment = bean.get("comment") or ""
                bean_parent = bean.get("parent") or ""
                if keyword_lower in bean["full_name"].lower():
                    match_score += 3
                if keyword_lower in bean_comment.lower():
                    match_score += 2
                for field in bean.get("fields", []):
                    if keyword_lower in field.get("name", "").lower():
                        match_score += 1

                if match_score > 0:
                    results.append({
                        "category": "bean",
                        "name": bean["full_name"],
                        "comment": bean_comment,
                        "match_score": match_score
                    })

        # 按匹配度排序
        results.sort(key=lambda x: x["match_score"], reverse=True)

        return {
            "keyword": keyword,
            "category": category,
            "count": len(results),
            "results": results[:20]  # 最多返回20个
        }

    def get_type_guide(self, topic: str = "all") -> Dict[str, Any]:
        """获取类型使用指南

        Args:
            topic: 指南主题 - basic(基本类型), container(容器类型), nullable(可空类型),
                   enum(枚举), bean(Bean), all(全部)

        Returns:
            类型使用指南
        """
        guides = {
            "basic": {
                "title": "基本类型指南",
                "description": "Luban 支持的基本数据类型",
                "types": [
                    {"type": "bool", "usage": "true/false", "note": "布尔值"},
                    {"type": "byte", "usage": "-128 ~ 127", "note": "小整数"},
                    {"type": "short", "usage": "-32768 ~ 32767", "note": "短整数"},
                    {"type": "int", "usage": "常用数字", "note": "32位整数，最常用"},
                    {"type": "long", "usage": "大整数", "note": "64位整数"},
                    {"type": "float", "usage": "3.14", "note": "单精度浮点数"},
                    {"type": "double", "usage": "高精度小数", "note": "双精度浮点数"},
                    {"type": "string", "usage": "文本内容", "note": "字符串"},
                    {"type": "text", "usage": "长文本", "note": "多行文本"},
                    {"type": "datetime", "usage": "2024-01-01", "note": "日期时间"}
                ]
            },
            "container": {
                "title": "容器类型指南",
                "description": "用于存储多个元素的数据结构",
                "types": [
                    {"type": "array<T>", "usage": "array<int>", "note": "定长数组，大小固定"},
                    {"type": "list<T>", "usage": "list<string>", "note": "变长列表，最常用"},
                    {"type": "set<T>", "usage": "set<int>", "note": "元素不重复"},
                    {"type": "map<K,V>", "usage": "map<int,string>", "note": "键值对映射"}
                ],
                "examples": [
                    {"scenario": "道具ID列表", "type": "list<int>", "data_example": "1001,1002,1003"},
                    {"scenario": "技能等级伤害", "type": "map<int,int>", "data_example": "1:100,2:200,3:300"},
                    {"scenario": "标签集合", "type": "set<string>", "data_example": "tag1,tag2,tag3"}
                ]
            },
            "nullable": {
                "title": "可空类型指南",
                "description": "在类型后加 ? 表示该字段可为空",
                "syntax": "Type? 或 基本类型?",
                "examples": [
                    {"type": "int?", "meaning": "整数或空", "use_case": "可选数值"},
                    {"type": "string?", "meaning": "字符串或空", "use_case": "可选描述"},
                    {"type": "list<int>?", "meaning": "列表或空", "use_case": "可选列表"}
                ],
                "note": "可空类型在数据表中可以留空，Luban 会将其解析为 null"
            },
            "enum": {
                "title": "枚举类型指南",
                "description": "枚举用于定义有限个命名常量",
                "definition": "在 __enums__.xlsx 中定义",
                "usage": [
                    {"step": 1, "action": "使用 enum list 查看所有枚举"},
                    {"step": 2, "action": "使用 enum get <name> 查看枚举详情"},
                    {"step": 3, "action": "在字段类型中使用枚举全名"}
                ],
                "example": {
                    "enum_name": "test.EQuality",
                    "values": "WHITE=0(白), GREEN=1(绿), BLUE=2(蓝)",
                    "field_usage": "quality:test.EQuality"
                }
            },
            "bean": {
                "title": "Bean 类型指南",
                "description": "Bean 是自定义数据结构，可以包含多个字段",
                "definition": "在 __beans__.xlsx 中定义",
                "usage": [
                    {"step": 1, "action": "使用 bean list 查看所有 Bean"},
                    {"step": 2, "action": "使用 bean get <name> 查看 Bean 详情"},
                    {"step": 3, "action": "在字段类型中使用 Bean 全名"}
                ],
                "example": {
                    "bean_name": "test.RewardItem",
                    "fields": "id:int, count:int",
                    "field_usage": "rewards:list<test.RewardItem>"
                }
            }
        }

        if topic == "all":
            return {
                "title": "Luban 类型系统完整指南",
                "topics": list(guides.keys()),
                "guides": guides
            }
        elif topic in guides:
            return guides[topic]
        else:
            return {"error": f"未知主题: {topic}", "available_topics": list(guides.keys())}

    def get_table_data(self, table_name: str) -> Optional[Dict[str, Any]]:
        """获取表的数据（从 Excel 文件）"""
        # 查找匹配的表定义
        tables = self.list_tables()
        target_table = None
        for t in tables:
            if t["full_name"] == table_name or t["full_name"].endswith("." + table_name):
                target_table = t
                break
        
        if not target_table:
            return None
        
        # 解析 input 字段找到对应的 Excel 文件
        input_val = target_table.get("input", "")
        if not input_val or isinstance(input_val, bool):
            # 如果是布尔值 true，尝试直接从 mode 字段找文件名
            mode_str = target_table.get("mode", "")
            if mode_str and isinstance(mode_str, str):
                excel_files = list(self.data_dir.glob(f"**/{mode_str}"))
                if excel_files:
                    try:
                        return self._parse_excel_data(excel_files[0])
                    except Exception as e:
                        print(f"读取 {excel_files[0]} 失败：{e}")
            return None
        
        # 简单的文件名提取（去掉路径和@符号后的内容）
        file_part = str(input_val).split("@")[0] if "@" in str(input_val) else str(input_val)
        excel_files = list(self.data_dir.glob("**/*.xlsx"))
        
        for excel_file in excel_files:
            if file_part in excel_file.as_posix() or file_part == excel_file.name:
                try:
                    return self._parse_excel_data(excel_file)
                except Exception as e:
                    print(f"读取 {excel_file} 失败：{e}")
        
        return None
    
    def _parse_excel_data(self, excel_file: Path, sheet_name: str = None) -> Dict[str, Any]:
        """解析 Excel 文件为数据结构
        
        Args:
            excel_file: Excel 文件路径
            sheet_name: Sheet 名（可选，默认使用第一个 sheet）
        """
        wb = openpyxl.load_workbook(excel_file)
        
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                print(f"错误: Sheet '{sheet_name}' 不存在")
                wb.close()
                return {"file": str(excel_file), "fields": [], "rows": []}
            sheet = wb[sheet_name]
        else:
            sheet = wb.active
        
        data = {
            "file": str(excel_file),
            "fields": [],
            "rows": []
        }
        
        # 解析字段定义
        var_row = None
        type_row = None
        desc_rows = []
        data_start_row = None
        
        for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if row[0] == "##var":
                var_row = row
            elif row[0] == "##type":
                type_row = row
            elif row[0] == "##":
                desc_rows.append(row)
            elif var_row and type_row:
                # 第一个非特殊行是数据开始
                data_start_row = i
                break
        
        if var_row and type_row:
            # 提取字段信息
            for j in range(1, len(var_row)):
                field_name = var_row[j] if j < len(var_row) else None
                field_type = type_row[j] if j < len(type_row) else None
                # 规范化: openpyxl 空单元格返回 None
                if field_type is None:
                    field_type = ""

                if field_name:
                    # 合并所有描述行的注释
                    comments = []
                    for desc_row in desc_rows:
                        if j < len(desc_row) and desc_row[j]:
                            comments.append(str(desc_row[j]))
                    
                    data["fields"].append({
                        "name": field_name,
                        "type": field_type,
                        "comment": " ".join(comments) if comments else ""
                    })
        
        # 解析数据行
        if data_start_row and data["fields"]:
            for i, row in enumerate(sheet.iter_rows(min_row=data_start_row, values_only=True), data_start_row):
                # 跳过空行
                if all(c is None for c in row):
                    continue
                
                row_data = {}
                for j, field in enumerate(data["fields"], 1):
                    value = row[j] if j < len(row) else None
                    if value is not None:
                        row_data[field["name"]] = value
                
                if row_data:
                    data["rows"].append(row_data)
        
        wb.close()
        return data
    
    # ==================== 缓存操作 ====================
    
    def get_file_hash(self, file_path: Path) -> str:
        """计算文件哈希"""
        if not file_path.exists():
            return ""
        
        with open(file_path, "rb") as f:
            return hashlib.md5(f.read()).hexdigest()
    
    def build_cache(self) -> bool:
        """构建缓存"""
        cache_file = self.cache_dir / "config_cache.json"
        
        cache_data = {
            "enums": self.list_enums(),
            "beans": self.list_beans(),
            "tables": self.list_tables(),
            "hashes": {
                "enums": self.get_file_hash(self.enums_file),
                "beans": self.get_file_hash(self.beans_file),
                "tables": self.get_file_hash(self.tables_file)
            }
        }
        
        with open(cache_file, "w", encoding="utf-8") as f:
            json.dump(cache_data, f, ensure_ascii=False, indent=2)
        
        print(f"✓ 已构建缓存: {cache_file}")
        return True
    
    def clear_cache(self) -> bool:
        """清除缓存"""
        import shutil
        if self.cache_dir.exists():
            shutil.rmtree(self.cache_dir)
            self.cache_dir.mkdir(exist_ok=True)
        print("✓ 已清除缓存")
        return True


def _parse_fields_arg(fields_str: str) -> list:
    """解析字段定义字符串，支持两种格式:
    1. JSON 数组: [{"name":"Id","type":"int","comment":"ID"}, ...]
    2. CSV 简写: name1:type1:comment1,name2:type2:comment2

    自动检测: 以 '[' 开头按 JSON 解析，否则按 CSV 解析。
    """
    fields_str = fields_str.strip()
    if not fields_str:
        return []

    # JSON 格式检测
    if fields_str.startswith("["):
        try:
            fields = json.loads(fields_str)
            # 验证每个字段至少有 name
            for f in fields:
                if not isinstance(f, dict) or "name" not in f:
                    raise ValueError(f"字段缺少 name: {f}")
            return fields
        except json.JSONDecodeError as e:
            print(f"错误: JSON 解析失败: {e}")
            return []

    # CSV 简写格式: name:type:comment,name:type:comment
    fields = []
    for field_str in fields_str.split(","):
        parts = field_str.split(":")
        if not parts[0].strip():
            continue
        fields.append({
            "name": parts[0].strip(),
            "type": parts[1].strip() if len(parts) > 1 else "",
            "comment": parts[2].strip() if len(parts) > 2 else ""
        })
    return fields


def main():
    parser = argparse.ArgumentParser(description="luban_skill - Luban 配置编辑器辅助脚本")
    parser.add_argument("--data-dir", default="DataTables/Datas", help="数据目录路径")
    
    subparsers = parser.add_subparsers(dest="command", help="命令")
    
    # 枚举命令
    enum_parser = subparsers.add_parser("enum", help="枚举操作")
    enum_subparsers = enum_parser.add_subparsers(dest="enum_command")
    
    enum_list = enum_subparsers.add_parser("list", help="列出所有枚举")
    enum_get = enum_subparsers.add_parser("get", help="获取枚举详情")
    enum_get.add_argument("name", metavar="NAME", help="枚举名称")
    enum_add = enum_subparsers.add_parser("add", help="新增枚举")
    enum_add.add_argument("name", metavar="NAME", help="枚举全名 (如 test.EWeaponType)")
    enum_add.add_argument("--values", required=True, help="枚举值，格式: name1=value1:alias1,name2=value2:alias2")
    enum_add.add_argument("--comment", default="", help="枚举注释")
    enum_add.add_argument("--flags", action="store_true", help="是否为标志枚举")
    enum_delete = enum_subparsers.add_parser("delete", help="删除枚举")
    enum_delete.add_argument("name", metavar="NAME", help="枚举名称")
    enum_delete.add_argument("--force", action="store_true", help="强制删除，忽略引用检查")
    
    # 更新枚举命令
    enum_update = enum_subparsers.add_parser("update", help="更新枚举属性")
    enum_update.add_argument("name", metavar="NAME", help="枚举名称")
    enum_update.add_argument("--comment", default=None, help="注释")
    enum_update.add_argument("--flags", default=None, action="store_true", help="是否为标志枚举")
    
    # Bean 命令
    bean_parser = subparsers.add_parser("bean", help="Bean 操作")
    bean_subparsers = bean_parser.add_subparsers(dest="bean_command")
    
    bean_list = bean_subparsers.add_parser("list", help="列出所有 Bean")
    bean_get = bean_subparsers.add_parser("get", help="获取 Bean 详情")
    bean_get.add_argument("name", metavar="NAME", help="Bean 名称")
    bean_add = bean_subparsers.add_parser("add", help="新增 Bean")
    bean_add.add_argument("name", metavar="NAME", help="Bean 全名")
    bean_add.add_argument("--fields", default="", help="字段定义，格式: name1:type1:comment1,name2:type2:comment2 或 JSON 数组")
    bean_add.add_argument("--file", default="", help="从 JSON 文件读取字段定义")
    bean_add.add_argument("--parent", default="", help="父类名称")
    bean_add.add_argument("--comment", default="", help="Bean 注释")
    bean_add.add_argument("--value-type", type=int, default=0, dest="value_type", help="是否为值类型（0=普通类，1=值类型/struct），用于 list<Bean> 中的 Bean 必须设为1")
    bean_add.add_argument("--sep", default="", help="分隔符（用于 list 类型元素分隔）")
    bean_delete = bean_subparsers.add_parser("delete", help="删除 Bean")
    bean_delete.add_argument("name", metavar="NAME", help="Bean 名称")
    bean_delete.add_argument("--force", action="store_true", help="强制删除，忽略引用检查")
    
    # 更新 Bean 命令
    bean_update = bean_subparsers.add_parser("update", help="更新 Bean 属性")
    bean_update.add_argument("name", metavar="NAME", help="Bean 名称")
    bean_update.add_argument("--sep", default=None, help="分隔符（用于 list 类型元素分隔，如 '|' 或 '#'")
    bean_update.add_argument("--comment", default=None, help="注释")
    bean_update.add_argument("--alias", default=None, help="别名")
    bean_update.add_argument("--parent", default=None, help="父类名称")
    bean_update.add_argument("--value-type", type=int, default=None, dest="value_type", help="是否为值类型（0=普通类，1=值类型/struct）")
    
    # 表命令
    table_parser = subparsers.add_parser("table", help="表操作")
    table_subparsers = table_parser.add_subparsers(dest="table_command")
    table_list = table_subparsers.add_parser("list", help="列出所有表")
    table_get = table_subparsers.add_parser("get", help="获取表数据")
    table_get.add_argument("name", metavar="NAME", help="表名称")
    
    # 新增表命令
    table_add = table_subparsers.add_parser("add", help="新增配置表")
    table_add.add_argument("name", metavar="NAME", help="表全名 (如 test.TbItem)")
    table_add.add_argument("--fields", required=True, 
                          help="字段定义，格式: name1:type1:comment1:group1,name2:type2:comment2:group2")
    table_add.add_argument("--value-type", default="", help="值类型")
    table_add.add_argument("--input", default="", help="输入文件名 (如 item.xlsx)")
    table_add.add_argument("--sheet", default="", help="Sheet名称（默认使用表名）")
    table_add.add_argument("--mode", default="", help="模式")
    table_add.add_argument("--comment", default="", help="表注释")
    table_add.add_argument("--index", default="", help="主键定义 (如 id 或 id1+id2)")
    table_add.add_argument("--groups", default="", help="分组列表 (如 c,s)，启用 ##group 行")
    table_add.add_argument("--auto-import", action="store_true", help="使用 # 前缀自动导入格式（不推荐，默认在 __tables__.xlsx 正式注册）")
    table_add.add_argument("--vertical", action="store_true", help="使用纵表模式（适合单例表）")
    
    # 删除表命令
    table_delete = table_subparsers.add_parser("delete", help="删除配置表")
    table_delete.add_argument("name", metavar="NAME", help="表名称")
    table_delete.add_argument("--delete-data", action="store_true", help="同时删除数据文件")
    
    # 更新表命令
    table_update = table_subparsers.add_parser("update", help="更新表属性")
    table_update.add_argument("name", metavar="NAME", help="表名称")
    table_update.add_argument("--comment", default=None, help="注释")
    table_update.add_argument("--input", default=None, help="输入文件名")
    table_update.add_argument("--mode", default=None, help="模式")
    table_update.add_argument("--value-type", default=None, help="值类型")
    
    # 检查老表命令
    table_check_legacy = table_subparsers.add_parser("check-legacy", help="检查可迁移到自动导入格式的表")
    
    # 迁移到自动导入格式命令
    table_migrate = table_subparsers.add_parser("migrate-auto", help="迁移表到自动导入格式")
    table_migrate.add_argument("name", metavar="NAME", nargs="?", default=None, help="表名称（不指定则迁移所有）")
    
    # 设置偏好命令
    pref_parser = subparsers.add_parser("pref", help="用户偏好设置")
    pref_subparsers = pref_parser.add_subparsers(dest="pref_command")
    
    pref_set = pref_subparsers.add_parser("set", help="设置偏好")
    pref_set.add_argument("key", metavar="KEY", help="偏好键名")
    pref_set.add_argument("value", metavar="VALUE", help="偏好值")
    
    pref_get = pref_subparsers.add_parser("get", help="获取偏好")
    pref_get.add_argument("key", metavar="KEY", help="偏好键名")
    
    pref_list = pref_subparsers.add_parser("list", help="列出所有偏好")
    
    # 字段操作命令
    field_parser = subparsers.add_parser("field", help="字段操作")
    field_subparsers = field_parser.add_subparsers(dest="field_command")
    
    # 列出字段
    field_list = field_subparsers.add_parser("list", help="列出表的所有字段")
    field_list.add_argument("table", metavar="TABLE", help="表名称")
    field_list.add_argument("--sheet", default="", help="Sheet名称（多 sheet 文件需要指定）")
    
    # 添加字段
    field_add = field_subparsers.add_parser("add", help="添加字段")
    field_add.add_argument("table", metavar="TABLE", help="表名称")
    field_add.add_argument("name", metavar="NAME", help="字段名")
    field_add.add_argument("--type", default="", help="字段类型")
    field_add.add_argument("--comment", default="", help="字段注释")
    field_add.add_argument("--group", default="", help="字段分组")
    field_add.add_argument("--sheet", default="", help="Sheet名称（多 sheet 文件需要指定）")
    field_add.add_argument("--position", type=int, default=-1, help="插入位置（从0开始，-1表示末尾）")
    
    # 修改字段
    field_update = field_subparsers.add_parser("update", help="修改字段")
    field_update.add_argument("table", metavar="TABLE", help="表名称")
    field_update.add_argument("name", metavar="NAME", help="原字段名")
    field_update.add_argument("--new-name", default=None, help="新字段名")
    field_update.add_argument("--type", default=None, help="新类型")
    field_update.add_argument("--comment", default=None, help="新注释")
    field_update.add_argument("--group", default=None, help="新分组")
    field_update.add_argument("--sheet", default="", help="Sheet名称（多 sheet 文件需要指定）")
    
    # 删除字段
    field_delete = field_subparsers.add_parser("delete", help="删除字段（危险操作）")
    field_delete.add_argument("table", metavar="TABLE", help="表名称")
    field_delete.add_argument("name", metavar="NAME", help="字段名")
    field_delete.add_argument("--sheet", default="", help="Sheet名称（多 sheet 文件需要指定）")
    field_delete.add_argument("--force", action="store_true", help="强制删除，跳过确认")
    
    # 禁用字段
    field_disable = field_subparsers.add_parser("disable", help="禁用字段（注释列，不导出但保留数据）")
    field_disable.add_argument("table", metavar="TABLE", help="表名称")
    field_disable.add_argument("name", metavar="NAME", help="字段名")
    field_disable.add_argument("--sheet", default="", help="Sheet名称（多 sheet 文件需要指定）")
    
    # 启用字段
    field_enable = field_subparsers.add_parser("enable", help="启用字段（取消注释列）")
    field_enable.add_argument("table", metavar="TABLE", help="表名称")
    field_enable.add_argument("name", metavar="NAME", help="字段名")
    field_enable.add_argument("--sheet", default="", help="Sheet名称（多 sheet 文件需要指定）")
    
    # 数据行操作命令
    row_parser = subparsers.add_parser("row", help="数据行操作")
    row_subparsers = row_parser.add_subparsers(dest="row_command")
    
    # 列出数据行
    row_list = row_subparsers.add_parser("list", help="列出数据行")
    row_list.add_argument("table", metavar="TABLE", help="表名称")
    row_list.add_argument("--sheet", default="", help="Sheet名称")
    row_list.add_argument("--start", type=int, default=0, help="起始行索引（从0开始）")
    row_list.add_argument("--limit", type=int, default=100, help="返回行数限制")
    
    # 添加数据行
    row_add = row_subparsers.add_parser("add", help="添加数据行")
    row_add.add_argument("table", metavar="TABLE", help="表名称")
    row_add.add_argument("--data", help="数据JSON格式，如 '{\"id\":1,\"name\":\"test\"}'")
    row_add.add_argument("--file", help="从JSON文件读取数据（推荐用于PowerShell）")
    row_add.add_argument("--sheet", default="", help="Sheet名称")
    
    # 更新数据行
    row_update = row_subparsers.add_parser("update", help="更新数据行")
    row_update.add_argument("table", metavar="TABLE", help="表名称")
    row_update.add_argument("index", metavar="INDEX", type=int, help="行索引（从0开始）")
    row_update.add_argument("--data", help="更新数据JSON格式")
    row_update.add_argument("--file", help="从JSON文件读取数据（推荐用于PowerShell）")
    row_update.add_argument("--sheet", default="", help="Sheet名称")
    
    # 删除数据行
    row_delete = row_subparsers.add_parser("delete", help="删除数据行")
    row_delete.add_argument("table", metavar="TABLE", help="表名称")
    row_delete.add_argument("index", metavar="INDEX", type=int, help="行索引（从0开始）")
    row_delete.add_argument("--sheet", default="", help="Sheet名称")
    row_delete.add_argument("--force", action="store_true", help="强制删除，跳过确认")
    
    # 查询单行（按字段值）
    row_get = row_subparsers.add_parser("get", help="按字段值查询数据行")
    row_get.add_argument("table", metavar="TABLE", help="表名称")
    row_get.add_argument("--field", required=True, help="字段名")
    row_get.add_argument("--value", required=True, help="字段值")
    row_get.add_argument("--sheet", default="", help="Sheet名称")
    
    # 多条件查询
    row_query = row_subparsers.add_parser("query", help="按多条件查询数据行")
    row_query.add_argument("table", metavar="TABLE", help="表名称")
    row_query.add_argument("--conditions", required=True, help="查询条件JSON，如 '{\"type\":\"Weapon\",\"quality\":5}'")
    row_query.add_argument("--sheet", default="", help="Sheet名称")
    row_query.add_argument("--limit", type=int, default=100, help="返回行数限制")
    
    # 批量操作命令
    batch_parser = subparsers.add_parser("batch", help="批量操作")
    batch_subparsers = batch_parser.add_subparsers(dest="batch_command")
    
    # 批量添加字段
    batch_fields = batch_subparsers.add_parser("fields", help="批量添加字段")
    batch_fields.add_argument("table", metavar="TABLE", help="表名称")
    batch_fields.add_argument("--data", required=True, help="字段JSON数组，如 '[{\"name\":\"f1\",\"type\":\"int\"}]'")
    batch_fields.add_argument("--sheet", default="", help="Sheet名称")
    
    # 批量添加数据行
    batch_rows = batch_subparsers.add_parser("rows", help="批量添加数据行")
    batch_rows.add_argument("table", metavar="TABLE", help="表名称")
    batch_rows.add_argument("--data", required=True, help="数据行JSON数组")
    batch_rows.add_argument("--sheet", default="", help="Sheet名称")
    
    # 导入导出命令
    export_parser = subparsers.add_parser("export", help="导出表数据为JSON")
    export_parser.add_argument("table", metavar="TABLE", help="表名称")
    export_parser.add_argument("--output", default=None, help="输出文件路径（默认打印到控制台）")
    export_parser.add_argument("--sheet", default="", help="Sheet名称")
    
    import_parser = subparsers.add_parser("import", help="从JSON导入数据")
    import_parser.add_argument("table", metavar="TABLE", help="表名称")
    import_parser.add_argument("file", metavar="FILE", help="输入JSON文件路径")
    import_parser.add_argument("--sheet", default="", help="Sheet名称")
    import_parser.add_argument("--mode", default="append", choices=["append", "replace"], help="导入模式")
    
    # 验证命令
    validate_parser = subparsers.add_parser("validate", help="验证表数据")
    validate_parser.add_argument("table", metavar="TABLE", nargs="?", default=None, help="表名称（不指定则验证所有表）")
    validate_parser.add_argument("--sheet", default="", help="Sheet名称")
    validate_parser.add_argument("--all", action="store_true", help="验证所有表")
    
    # Luban CLI 命令
    gen_parser = subparsers.add_parser("gen", help="调用 Luban CLI 生成代码")
    gen_parser.add_argument("--output", default=None, help="输出目录")
    gen_parser.add_argument("--luban-cmd", default="dotnet run --project Luban.CLI", help="Luban CLI 命令")
    
    # 引用检查命令
    ref_parser = subparsers.add_parser("ref", help="引用完整性检查")
    ref_parser.add_argument("type", metavar="TYPE", help="类型名称（枚举或Bean）")
    
    # 模板命令
    template_parser = subparsers.add_parser("template", help="配置模板操作")
    template_subparsers = template_parser.add_subparsers(dest="template_command")
    
    template_list = template_subparsers.add_parser("list", help="列出所有模板")
    template_create = template_subparsers.add_parser("create", help="从模板创建表")
    template_create.add_argument("template", metavar="TEMPLATE", help="模板名称")
    template_create.add_argument("table", metavar="TABLE", help="表名称（不含模块）")
    template_create.add_argument("--module", default="test", help="模块名")
    template_create.add_argument("--input", default=None, help="输入文件名")
    
    # 迁移命令
    rename_parser = subparsers.add_parser("rename", help="重命名表")
    rename_parser.add_argument("old_name", metavar="OLD_NAME", help="原表名")
    rename_parser.add_argument("new_name", metavar="NEW_NAME", help="新表名")
    rename_parser.add_argument("--migrate-data", action="store_true", help="迁移数据文件")
    
    copy_parser = subparsers.add_parser("copy", help="复制表")
    copy_parser.add_argument("source", metavar="SOURCE", help="源表名")
    copy_parser.add_argument("target", metavar="TARGET", help="目标表名")
    copy_parser.add_argument("--copy-data", action="store_true", help="复制数据文件")
    
    # 差异对比命令
    diff_parser = subparsers.add_parser("diff", help="差异对比")
    diff_parser.add_argument("table1", metavar="TABLE1", help="表1名称")
    diff_parser.add_argument("table2", metavar="TABLE2", help="表2名称或JSON文件")
    diff_parser.add_argument("--json", action="store_true", help="table2 是 JSON 文件路径")
    
    # 自动导入表命令
    auto_parser = subparsers.add_parser("auto", help="自动导入表操作")
    auto_subparsers = auto_parser.add_subparsers(dest="auto_command")
    
    auto_list = auto_subparsers.add_parser("list", help="列出自动导入的表")
    auto_create = auto_subparsers.add_parser("create", help="创建自动导入表")
    auto_create.add_argument("name", metavar="NAME", help="表名（如 #Item 或 #Item-道具表）")
    auto_create.add_argument("--fields", default="", help="字段定义，格式 name:type:comment")
    
    # 常量别名命令
    alias_parser = subparsers.add_parser("alias", help="常量别名操作")
    alias_subparsers = alias_parser.add_subparsers(dest="alias_command")
    
    alias_list = alias_subparsers.add_parser("list", help="列出所有常量别名")
    alias_add = alias_subparsers.add_parser("add", help="添加常量别名")
    alias_add.add_argument("name", metavar="NAME", help="别名名")
    alias_add.add_argument("value", metavar="VALUE", help="别名值")
    alias_add.add_argument("--comment", default="", help="注释")
    alias_delete = alias_subparsers.add_parser("delete", help="删除常量别名")
    alias_delete.add_argument("name", metavar="NAME", help="别名名")
    alias_resolve = alias_subparsers.add_parser("resolve", help="解析常量别名")
    alias_resolve.add_argument("name", metavar="NAME", help="别名名")
    
    # 数据标签命令
    tag_parser = subparsers.add_parser("tag", help="数据标签操作")
    tag_subparsers = tag_parser.add_subparsers(dest="tag_command")
    
    tag_list = tag_subparsers.add_parser("list", help="列出表的数据标签")
    tag_list.add_argument("table", metavar="TABLE", help="表名称")
    tag_list.add_argument("--sheet", default="", help="Sheet名称")
    
    tag_add = tag_subparsers.add_parser("add", help="给数据行添加标签")
    tag_add.add_argument("table", metavar="TABLE", help="表名称")
    tag_add.add_argument("index", metavar="INDEX", type=int, help="行索引")
    tag_add.add_argument("tag", metavar="TAG", help="标签名")
    tag_add.add_argument("--sheet", default="", help="Sheet名称")
    
    tag_remove = tag_subparsers.add_parser("remove", help="移除数据行标签")
    tag_remove.add_argument("table", metavar="TABLE", help="表名称")
    tag_remove.add_argument("index", metavar="INDEX", type=int, help="行索引")
    tag_remove.add_argument("--sheet", default="", help="Sheet名称")
    
    # 字段变体命令
    variant_parser = subparsers.add_parser("variant", help="字段变体操作")
    variant_subparsers = variant_parser.add_subparsers(dest="variant_command")
    
    variant_list = variant_subparsers.add_parser("list", help="列出字段变体")
    variant_list.add_argument("table", metavar="TABLE", help="表名称")
    variant_list.add_argument("field", metavar="FIELD", help="字段名")
    variant_list.add_argument("--sheet", default="", help="Sheet名称")
    
    variant_add = variant_subparsers.add_parser("add", help="添加字段变体")
    variant_add.add_argument("table", metavar="TABLE", help="表名称")
    variant_add.add_argument("field", metavar="FIELD", help="字段名")
    variant_add.add_argument("variant", metavar="VARIANT", help="变体名（如 zh, en）")
    variant_add.add_argument("--sheet", default="", help="Sheet名称")
    
    # 多行结构命令
    multirow_parser = subparsers.add_parser("multirow", help="多行结构列表操作")
    multirow_parser.add_argument("table", metavar="TABLE", help="表名称")
    multirow_parser.add_argument("field", metavar="FIELD", help="字段名")
    multirow_parser.add_argument("--disable", action="store_true", help="禁用多行结构")
    multirow_parser.add_argument("--sheet", default="", help="Sheet名称")
    
    # 类型信息命令
    type_parser = subparsers.add_parser("type", help="类型系统操作")
    type_subparsers = type_parser.add_subparsers(dest="type_command")

    # type info - 查询单个类型详情
    type_info = type_subparsers.add_parser("info", help="查询类型详情")
    type_info.add_argument("name", metavar="NAME", help="类型名，如 'int', 'list<int>', 'test.EQuality'")

    # type list - 列出所有可用类型
    type_list = type_subparsers.add_parser("list", help="列出所有可用类型")
    type_list.add_argument("--category", choices=["basic", "container", "enum", "bean", "all"],
                          default="all", help="类型类别筛选")

    # type validate - 验证类型是否有效
    type_validate = type_subparsers.add_parser("validate", help="验证类型字符串是否有效")
    type_validate.add_argument("name", metavar="NAME", help="类型名字符串")

    # type suggest - 根据字段名建议类型
    type_suggest = type_subparsers.add_parser("suggest", help="根据字段名建议类型")
    type_suggest.add_argument("field_name", metavar="FIELD_NAME", help="字段名，如 'item_id', 'name', 'quality'")
    type_suggest.add_argument("--context", choices=["item", "skill", "monster", "quest", "general"],
                             default="general", help="上下文类型")

    # type search - 搜索类型
    type_search = type_subparsers.add_parser("search", help="搜索类型")
    type_search.add_argument("keyword", metavar="KEYWORD", help="搜索关键词")
    type_search.add_argument("--category", choices=["enum", "bean", "all"],
                            default="all", help="搜索类别")

    # type guide - 显示类型使用指南
    type_guide = type_subparsers.add_parser("guide", help="显示类型使用指南")
    type_guide.add_argument("--topic", choices=["basic", "container", "nullable", "enum", "bean", "all"],
                           default="all", help="指南主题")
    
    # 缓存命令
    cache_parser = subparsers.add_parser("cache", help="缓存操作")
    cache_subparsers = cache_parser.add_subparsers(dest="cache_command")
    cache_build = cache_subparsers.add_parser("build", help="构建缓存")
    cache_clear = cache_subparsers.add_parser("clear", help="清除缓存")
    
    args = parser.parse_args()
    
    if not args.command:
        parser.print_help()
        return
    
    helper = LubanConfigHelper(args.data_dir)
    
    # 枚举操作
    if args.command == "enum":
        if args.enum_command == "list":
            enums = helper.list_enums()
            print(json.dumps(enums, ensure_ascii=False, indent=2))
        elif args.enum_command == "get":
            enum = helper.get_enum(args.name)
            if enum:
                print(json.dumps(enum, ensure_ascii=False, indent=2))
            else:
                print(f"未找到枚举: {args.name}")
        elif args.enum_command == "add":
            # 解析枚举值
            items = []
            for item_str in args.values.split(","):
                parts = item_str.split("=")
                name = parts[0]
                value_alias = parts[1].split(":") if len(parts) > 1 else ["", ""]
                value = value_alias[0] if value_alias[0].isdigit() else None
                alias = value_alias[1] if len(value_alias) > 1 else ""
                if value is None:
                    value = len(items) + 1
                items.append({"name": name, "value": int(value), "alias": alias})
            helper.add_enum(args.name, items, args.flags, True, args.comment)
        elif args.enum_command == "delete":
            helper.delete_enum_safe(args.name, args.force)
        elif args.enum_command == "update":
            helper.update_enum(
                enum_name=args.name,
                comment=args.comment,
                flags=args.flags
            )
    
    # Bean 操作
    elif args.command == "bean":
        if args.bean_command == "list":
            beans = helper.list_beans()
            print(json.dumps(beans, ensure_ascii=False, indent=2))
        elif args.bean_command == "get":
            bean = helper.get_bean(args.name)
            if bean:
                print(json.dumps(bean, ensure_ascii=False, indent=2))
            else:
                print(f"未找到 Bean: {args.name}")
        elif args.bean_command == "add":
            # 解析字段: 支持 --fields (CSV或JSON) 或 --file (JSON文件)
            fields_str = args.fields
            if args.file:
                with open(args.file, "r", encoding="utf-8") as f:
                    fields_str = f.read().strip()
            if not fields_str:
                print("错误: 需要指定 --fields 或 --file")
                return
            fields = _parse_fields_arg(fields_str)
            helper.add_bean(args.name, fields, args.parent, args.comment,
                            value_type=args.value_type, sep=args.sep)
        elif args.bean_command == "delete":
            helper.delete_bean_safe(args.name, args.force)
        elif args.bean_command == "update":
            helper.update_bean(
                bean_name=args.name,
                sep=args.sep,
                comment=args.comment,
                alias=args.alias,
                parent=args.parent,
                value_type=args.value_type
            )
    
    # 表操作
    elif args.command == "table":
        if args.table_command == "list":
            tables = helper.list_tables()
            print(json.dumps(tables, ensure_ascii=False, indent=2))
        elif args.table_command == "get":
            table_data = helper.get_table_data(args.name)
            if table_data:
                print(json.dumps(table_data, ensure_ascii=False, indent=2))
            else:
                print(f"未找到表：{args.name}")
        elif args.table_command == "add":
            # 解析字段: 支持 JSON 数组或 CSV 简写
            raw_fields = _parse_fields_arg(args.fields)
            fields = []
            for f in raw_fields:
                field_name = f.get("name", "")
                field_type = f.get("type", "")
                field_comment = f.get("comment", "")
                field_group = f.get("group", "")

                # 自动推断分组
                if not field_group:
                    field_group = helper._infer_field_group(field_name, field_type)

                field = {
                    "name": field_name,
                    "type": field_type,
                    "comment": field_comment,
                    "group": field_group
                }
                fields.append(field)
            
            # 解析分组
            groups = args.groups.split(",") if args.groups else None
            
            # 判断是否使用自动导入格式
            auto_import = args.auto_import if hasattr(args, 'auto_import') else None
            vertical = args.vertical if hasattr(args, 'vertical') else False
            
            helper.add_table(
                full_name=args.name,
                fields=fields,
                value_type=args.value_type,
                input_file=args.input,
                sheet_name=args.sheet,
                mode=args.mode,
                comment=args.comment,
                index=args.index,
                groups=groups,
                auto_import=auto_import,
                vertical=vertical
            )
        elif args.table_command == "delete":
            helper.delete_table(args.name, args.delete_data)
        elif args.table_command == "update":
            helper.update_table(
                table_name=args.name,
                comment=args.comment,
                input_file=args.input,
                mode=args.mode,
                value_type=args.value_type
            )
        elif args.table_command == "check-legacy":
            helper.suggest_auto_import_migration()
        elif args.table_command == "migrate-auto":
            helper.migrate_to_auto_import(args.name if hasattr(args, 'name') else None)
    
    # 偏好设置操作
    elif args.command == "pref":
        if args.pref_command == "set":
            # 解析值类型
            value = args.value
            if value.lower() == "true":
                value = True
            elif value.lower() == "false":
                value = False
            helper.set_preference(args.key, value)
        elif args.pref_command == "get":
            value = helper.get_preference(args.key)
            print(f"{args.key} = {value}")
        elif args.pref_command == "list":
            print(json.dumps(helper.USER_PREFERENCES, ensure_ascii=False, indent=2))
    
    # 字段操作
    elif args.command == "field":
        sheet_name = args.sheet if args.sheet else None
        
        if args.field_command == "list":
            fields = helper.list_fields(args.table, sheet_name)
            if fields:
                print(json.dumps(fields, ensure_ascii=False, indent=2))
            else:
                print(f"未找到表或表无字段: {args.table}")
        
        elif args.field_command == "add":
            helper.add_field(
                table_name=args.table,
                field_name=args.name,
                field_type=args.type,
                field_comment=args.comment,
                field_group=args.group,
                sheet_name=sheet_name,
                position=args.position
            )
        
        elif args.field_command == "update":
            helper.update_field(
                table_name=args.table,
                field_name=args.name,
                new_name=args.new_name,
                new_type=args.type,
                new_comment=args.comment,
                new_group=args.group,
                sheet_name=sheet_name
            )
        
        elif args.field_command == "delete":
            helper.delete_field(
                table_name=args.table,
                field_name=args.name,
                sheet_name=sheet_name,
                force=args.force
            )
        
        elif args.field_command == "disable":
            helper.disable_field(
                table_name=args.table,
                field_name=args.name,
                sheet_name=sheet_name
            )
        
        elif args.field_command == "enable":
            helper.enable_field(
                table_name=args.table,
                field_name=args.name,
                sheet_name=sheet_name
            )
    
    # 数据行操作
    elif args.command == "row":
        import json as json_module
        sheet_name = args.sheet if args.sheet else None
        
        if args.row_command == "list":
            rows = helper.list_rows(
                table_name=args.table,
                sheet_name=sheet_name,
                start=args.start,
                limit=args.limit
            )
            if rows is not None:
                print(json.dumps(rows, ensure_ascii=False, indent=2))
        
        elif args.row_command == "add":
            # 支持从文件或命令行读取JSON
            json_str = None
            if args.file:
                try:
                    with open(args.file, 'r', encoding='utf-8') as f:
                        json_str = f.read()
                except FileNotFoundError:
                    print(f"错误: 文件不存在 - {args.file}")
                    return
            elif args.data:
                json_str = args.data
            else:
                print("错误: 需要指定 --data 或 --file")
                return
            
            try:
                data = json_module.loads(json_str)
                helper.add_row(
                    table_name=args.table,
                    data=data,
                    sheet_name=sheet_name
                )
            except json_module.JSONDecodeError as e:
                print(f"错误: JSON 格式无效 - {e}")
        
        elif args.row_command == "update":
            # 支持从文件或命令行读取JSON
            json_str = None
            if args.file:
                try:
                    with open(args.file, 'r', encoding='utf-8') as f:
                        json_str = f.read()
                except FileNotFoundError:
                    print(f"错误: 文件不存在 - {args.file}")
                    return
            elif args.data:
                json_str = args.data
            else:
                print("错误: 需要指定 --data 或 --file")
                return
            
            try:
                data = json_module.loads(json_str)
                helper.update_row(
                    table_name=args.table,
                    row_index=args.index,
                    data=data,
                    sheet_name=sheet_name
                )
            except json_module.JSONDecodeError as e:
                print(f"错误: JSON 格式无效 - {e}")
        
        elif args.row_command == "delete":
            helper.delete_row(
                table_name=args.table,
                row_index=args.index,
                sheet_name=sheet_name,
                force=args.force
            )
        elif args.row_command == "get":
            result = helper.get_row(
                table_name=args.table,
                field=args.field,
                value=args.value,
                sheet_name=sheet_name
            )
            if result:
                print(json.dumps(result, ensure_ascii=False, indent=2))
        elif args.row_command == "query":
            import json as json_module
            try:
                conditions = json_module.loads(args.conditions)
            except json_module.JSONDecodeError:
                print("错误: conditions 不是有效的 JSON 格式")
                sys.exit(1)
            result = helper.query_rows(
                table_name=args.table,
                conditions=conditions,
                sheet_name=sheet_name,
                limit=args.limit
            )
            if result:
                print(json.dumps(result, ensure_ascii=False, indent=2))
    
    # 批量操作
    elif args.command == "batch":
        import json as json_module
        sheet_name = args.sheet if args.sheet else None
        
        if args.batch_command == "fields":
            try:
                fields = json_module.loads(args.data)
                helper.batch_add_fields(
                    table_name=args.table,
                    fields=fields,
                    sheet_name=sheet_name
                )
            except json_module.JSONDecodeError as e:
                print(f"错误: JSON 格式无效 - {e}")
        
        elif args.batch_command == "rows":
            try:
                rows = json_module.loads(args.data)
                helper.batch_add_rows(
                    table_name=args.table,
                    rows=rows,
                    sheet_name=sheet_name
                )
            except json_module.JSONDecodeError as e:
                print(f"错误: JSON 格式无效 - {e}")
    
    # 导入导出操作
    elif args.command == "export":
        sheet_name = args.sheet if args.sheet else None
        helper.export_json(
            table_name=args.table,
            output_file=args.output,
            sheet_name=sheet_name
        )
    
    elif args.command == "import":
        sheet_name = args.sheet if args.sheet else None
        helper.import_json(
            table_name=args.table,
            input_file=args.file,
            sheet_name=sheet_name,
            mode=args.mode
        )
    
    # 验证操作
    elif args.command == "validate":
        if args.all or not args.table:
            # 验证所有表
            results = helper.validate_all()
            print(json.dumps(results, ensure_ascii=False, indent=2))
        else:
            # 验证单个表
            sheet_name = args.sheet if args.sheet else None
            result = helper.validate_table(args.table, sheet_name)
            print(json.dumps(result, ensure_ascii=False, indent=2))
    
    # Luban CLI 操作
    elif args.command == "gen":
        helper.gen(
            output_dir=args.output,
            luban_cmd=args.luban_cmd
        )
    
    # 引用检查操作
    elif args.command == "ref":
        result = helper.check_references(args.type)
        print(json.dumps(result, ensure_ascii=False, indent=2))
    
    # 模板操作
    elif args.command == "template":
        if args.template_command == "list":
            templates = helper.list_templates()
            print(json.dumps(templates, ensure_ascii=False, indent=2))
        elif args.template_command == "create":
            helper.create_from_template(
                template_name=args.template,
                table_name=args.table,
                input_file=args.input,
                module=args.module
            )
    
    # 迁移操作
    elif args.command == "rename":
        helper.rename_table(
            old_name=args.old_name,
            new_name=args.new_name,
            migrate_data=args.migrate_data
        )
    
    elif args.command == "copy":
        helper.copy_table(
            source_name=args.source,
            target_name=args.target,
            copy_data=args.copy_data
        )
    
    # 差异对比操作
    elif args.command == "diff":
        if args.json:
            result = helper.diff_with_json(args.table1, args.table2)
        else:
            result = helper.diff_tables(args.table1, args.table2)
        print(json.dumps(result, ensure_ascii=False, indent=2))
    
    # 自动导入表操作
    elif args.command == "auto":
        if args.auto_command == "list":
            tables = helper.list_auto_import_tables()
            print(json.dumps(tables, ensure_ascii=False, indent=2))
        elif args.auto_command == "create":
            helper.create_auto_import_table(args.name, args.fields)
    
    # 常量别名操作
    elif args.command == "alias":
        if args.alias_command == "list":
            aliases = helper.list_constalias()
            print(json.dumps(aliases, ensure_ascii=False, indent=2))
        elif args.alias_command == "add":
            helper.add_constalias(args.name, args.value, args.comment)
        elif args.alias_command == "delete":
            helper.delete_constalias(args.name)
        elif args.alias_command == "resolve":
            value = helper.resolve_constalias(args.name)
            if value:
                print(f"{args.name} = {value}")
            else:
                print(f"未找到常量别名: {args.name}")
    
    # 数据标签操作
    elif args.command == "tag":
        sheet_name = args.sheet if hasattr(args, 'sheet') and args.sheet else None
        
        if args.tag_command == "list":
            result = helper.list_data_tags(args.table, sheet_name)
            print(json.dumps(result, ensure_ascii=False, indent=2))
        elif args.tag_command == "add":
            helper.tag_row(args.table, args.index, args.tag, sheet_name)
        elif args.tag_command == "remove":
            helper.untag_row(args.table, args.index, sheet_name)
    
    # 字段变体操作
    elif args.command == "variant":
        sheet_name = args.sheet if hasattr(args, 'sheet') and args.sheet else None
        
        if args.variant_command == "list":
            result = helper.list_field_variants(args.table, args.field, sheet_name)
            print(json.dumps(result, ensure_ascii=False, indent=2))
        elif args.variant_command == "add":
            helper.add_field_variant(args.table, args.field, args.variant, sheet_name)
    
    # 多行结构操作
    elif args.command == "multirow":
        sheet_name = args.sheet if args.sheet else None
        enable = not args.disable
        helper.set_multi_row_field(args.table, args.field, enable, sheet_name)
    
    # 类型信息操作
    elif args.command == "type":
        if not args.type_command or args.type_command == "info":
            # 默认查询类型详情
            type_name = args.name if hasattr(args, 'name') else None
            if type_name:
                result = helper.get_type_info(type_name)
                print(json.dumps(result, ensure_ascii=False, indent=2))
            else:
                print("错误: 请指定类型名或使用子命令")
        elif args.type_command == "list":
            result = helper.list_all_types(args.category)
            print(json.dumps(result, ensure_ascii=False, indent=2))
        elif args.type_command == "validate":
            result = helper.validate_type(args.name)
            print(json.dumps(result, ensure_ascii=False, indent=2))
        elif args.type_command == "suggest":
            result = helper.suggest_type_for_field(args.field_name, args.context)
            print(json.dumps(result, ensure_ascii=False, indent=2))
        elif args.type_command == "search":
            result = helper.search_types(args.keyword, args.category)
            print(json.dumps(result, ensure_ascii=False, indent=2))
        elif args.type_command == "guide":
            result = helper.get_type_guide(args.topic)
            print(json.dumps(result, ensure_ascii=False, indent=2))
    
    # 缓存操作
    elif args.command == "cache":
        if args.cache_command == "build":
            helper.build_cache()
        elif args.cache_command == "clear":
            helper.clear_cache()


if __name__ == "__main__":
    main()
